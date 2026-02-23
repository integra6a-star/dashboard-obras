import json
from pathlib import Path
from datetime import datetime, date
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
import re
import unicodedata

# =========================
# CAMINHOS (ROBUSTO)
# =========================
SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent

EXCEL_NAME = "BASE_DASH_EXTENSAO_POWERBI.xlsx"

CANDIDATOS_EXCEL = [
    ROOT_DIR / "docs" / EXCEL_NAME,
    ROOT_DIR / EXCEL_NAME,
    SCRIPT_DIR / EXCEL_NAME,
]

ARQ_EXCEL = next((p for p in CANDIDATOS_EXCEL if p.exists()), None)
SAIDA_JSON = (ROOT_DIR / "docs" / "dados.json").resolve()

# =========================
# MAPEAMENTO DE COLUNAS
# =========================
REQUIRED = {
    "Data": ["data"],
    "Obra": ["obra"],
    "Bloco": ["bloco"],
    "Tipo": ["tipo_extensao", "tipo extensao", "tipo", "tipoextensao"],
    "Planejado_m": ["extensao_planejada_m", "extensao planejada (m)", "extensao planejada", "planejado_m"],
    "Executado_m": ["extensao_executada_m", "extensao executada (m)", "extensao executada", "executado_m"],
}

OPTIONAL = {
    "PV": ["pv", "pvs", "qtd pv", "quantidade pv", "qtdpv"],
    "Profundidade_m": ["profundidade_pv_m", "profundidade pv (m)", "profundidade", "profundidade_m"],
    "Economias_Previstas": ["economias prevista", "economias previstas", "econ prev", "economias prevista(s)"],
    "Economias_Recebidas": ["economias recebidas", "economias recebida", "econ receb", "economias recebida(s)"],
}

def strip_accents(s: str) -> str:
    if s is None:
        return ""
    return unicodedata.normalize("NFD", str(s)).encode("ascii", "ignore").decode("ascii")

MESES_PT = {
    "janeiro":1, "jan":1,
    "fevereiro":2, "fev":2,
    "marco":3, "mar":3,
    "abril":4, "abr":4,
    "maio":5, "mai":5,
    "junho":6, "jun":6,
    "julho":7, "jul":7,
    "agosto":8, "ago":8,
    "setembro":9, "set":9, "sete":9,
    "outubro":10, "out":10,
    "novembro":11, "nov":11,
    "dezembro":12, "dez":12,
}

def detect_producao_cols(headers):
    """Detecta colunas do tipo 'produção março 2025' e retorna dict {YYYY-MM: idx}."""
    cols = {}
    for idx, h in enumerate(headers):
        if h is None:
            continue
        hs = strip_accents(h).lower().strip()
        if not hs.startswith("producao"):
            continue
        rest = hs[len("producao"):].strip()
        rest = re.sub(r"^[\s:\-_/]+", "", rest)

        # Ex: "03 2025" ou "03/2025"
        m = re.search(r"(?P<mes>\d{1,2})\s*[\/\-\s]\s*(?P<ano>\d{4})", rest)
        if m:
            mes = int(m.group("mes"))
            ano = int(m.group("ano"))
            if 1 <= mes <= 12:
                key = f"{ano:04d}-{mes:02d}"
                cols.setdefault(key, idx)
            continue

        # Ex: "marco 2025"
        m = re.search(r"(?P<mes>[a-z]+)\s+(?P<ano>\d{4})", rest)
        if m:
            mes_txt = m.group("mes").strip()
            ano = int(m.group("ano"))
            mes = MESES_PT.get(mes_txt)
            if mes:
                key = f"{ano:04d}-{mes:02d}"
                cols.setdefault(key, idx)
    return cols

def norm(s):
    return "".join(str(s).strip().lower().split())

def to_float(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except:
        return 0.0

def to_iso_date(v):
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()

    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except:
            pass
    return ""

def find_col(headers_norm, candidates):
    for cand in candidates:
        cn = norm(cand)
        for i, hn in enumerate(headers_norm):
            if hn == cn:
                return i
    return None

def main():
    if ARQ_EXCEL is None:
        raise FileNotFoundError(
            f"Não encontrei o Excel '{EXCEL_NAME}'. Coloque ele na raiz ou em docs/."
        )

    wb = load_workbook(ARQ_EXCEL, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in ws[1]]
    if not headers or all(h is None for h in headers):
        raise ValueError("Primeira linha da planilha está vazia (sem cabeçalhos).")

    headers_norm = [norm(h) if h is not None else "" for h in headers]

    # Produção mensal (colunas do tipo 'produção março 2025')
    prod_cols = detect_producao_cols(headers)
    meses_producao = sorted(prod_cols.keys())

    col_index = {}
    # required
    for out_name, candidates in REQUIRED.items():
        found = find_col(headers_norm, candidates)
        if found is None:
            raise ValueError(
                f"Não encontrei a coluna obrigatória '{out_name}'.\n"
                f"Cabeçalhos encontrados: {headers}"
            )
        col_index[out_name] = found

    # optional
    missing_optional = []
    for out_name, candidates in OPTIONAL.items():
        found = find_col(headers_norm, candidates)
        if found is None:
            missing_optional.append(out_name)
        else:
            col_index[out_name] = found

    registros = []
    datas_vazias = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue

        d_iso = to_iso_date(row[col_index["Data"]])
        if not d_iso:
            datas_vazias += 1

        def opt(name):
            idx = col_index.get(name, None)
            return to_float(row[idx]) if idx is not None else 0.0

        registros.append({
            "Data": d_iso,
            "Obra": str(row[col_index["Obra"]] or "").strip(),
            "Bloco": str(row[col_index["Bloco"]] or "").strip(),
            "Tipo": str(row[col_index["Tipo"]] or "").strip(),
            "Planejado_m": to_float(row[col_index["Planejado_m"]]),
            "Executado_m": to_float(row[col_index["Executado_m"]]),
            "PV": opt("PV"),
            "Profundidade_m": opt("Profundidade_m"),
            "Economias_Previstas": opt("Economias_Previstas"),
            "Economias_Recebidas": opt("Economias_Recebidas"),
            "ProducaoMensal": {k: to_float(row[idx]) for k, idx in prod_cols.items() if idx is not None and row[idx] not in (None, "")}
        })

    payload = {
        "atualizado_em": datetime.now(ZoneInfo("America/Sao_Paulo")).isoformat(),
        "meses_producao": meses_producao,
        "registros": registros
    }

    SAIDA_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(SAIDA_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"OK: {SAIDA_JSON} gerado com {len(registros)} linhas.")
    print(f"Excel lido de: {ARQ_EXCEL}")
    if not meses_producao:
        print("Aviso: nenhuma coluna de 'produção <mês> <ano>' foi encontrada.")
    if missing_optional:
        print("Aviso: colunas opcionais ausentes (preenchidas com 0): " + ", ".join(missing_optional))
    if datas_vazias > 0:
        print(f"Aviso: {datas_vazias} linhas ficaram com Data vazia (Curva S ignora).")

if __name__ == "__main__":
    main()