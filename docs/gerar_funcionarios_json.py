import json
from datetime import datetime
import re
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ========= CONFIG (ROBUSTO) =========
SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent

# Nomes candidatos (porque a planilha muda e às vezes tem espaço/underscore diferente)
EXCEL_CANDIDATOS_NOME = [
    "02.26_Estrutura Integra 6A_cleulton fevereiro.xlsx",   # seu nome atual (print)
    "02.26_Estrutura Integra 6A _cleulton fevereiro.xlsx",  # seu nome antigo (com espaço)
    "02.26_Estrutura Integra 6A _cleulton fevererio.xlsx",  # possível typo antigo
]

# Procura em docs/ e na raiz
CANDIDATOS_EXCEL = []
for n in EXCEL_CANDIDATOS_NOME:
    CANDIDATOS_EXCEL += [
        ROOT_DIR / "docs" / n,
        ROOT_DIR / n,
        SCRIPT_DIR / n,
    ]

ARQ_EXCEL = next((p for p in CANDIDATOS_EXCEL if p.exists()), None)

# Saídas (sempre em docs/)
ARQUIVO_SAIDA = (ROOT_DIR / "docs" / "funcionarios.json").resolve()
ARQUIVO_HIST = (ROOT_DIR / "docs" / "funcionarios_historico.json").resolve()

# Aba preferida. Se não existir, tenta achar automaticamente.
ABA_PREFERIDA = "Atualizado"

# Se seu cabeçalho estiver em outra linha, ajuste aqui.
LINHA_CABECALHO_FIXA = None  # ex: 1 ou 2. None = autodetect

# Fallback por letra (se o cabeçalho mudar)
COL_FIXA = {
    "Setor": "A",
    "Equipe": "B",
    "Custo": "C",
    "Matricula": "E",
    "Nome": "F",
    "Funcao": "G",
    "Admissao": "H",
    "Rescisao": "I",
    "Status": "J",
    "Horario": "K",
    "Regime": "L",
    "Salario": "M",
    "ValeRefeicao": "N",
    "ValorVeiculo": "O",
    "Combustivel": "P",
    "Veiculo": "Q",
    "Placa": "R",
}
# ===================================

TZ = ZoneInfo("America/Sao_Paulo")


def now_iso_local():
    return datetime.now(TZ).isoformat()


def now_iso_utc():
    return datetime.now(ZoneInfo("UTC")).isoformat()


def norm_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()

    mapa = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    for k, v in mapa.items():
        s = s.replace(k, v)

    s = re.sub(r"\s+", " ", s)
    return s


def to_float(v):
    """Converte moeda/texto/número para float."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()
    if s == "" or s == "-":
        return 0.0

    # remove R$, espaços e tudo que não seja dígito/ponto/vírgula/sinal
    s = re.sub(r"[^\d,.\-]", "", s)

    # BR: 2.867,00 -> 2867.00
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")

    try:
        return float(s)
    except:
        return 0.0


def to_text(v):
    if v is None:
        return ""
    return str(v).strip()


def infer_mes_competencia():
    # usa mês atual (do seu histórico)
    now = datetime.now(TZ)
    return f"{now.year:04d}-{now.month:02d}"


def detect_tipo(regime_txt: str) -> str:
    s = norm_header(regime_txt)
    if "pj" in s or "cnpj" in s or "jurid" in s:
        return "PJ"
    if "clt" in s:
        return "CLT"
    return ""


def normalize_status(status_txt: str) -> str:
    s = norm_header(status_txt)
    if s in ("", "ativo", "ativa", "ok", "trabalhando"):
        return "Ativo"
    if "afast" in s or "ferias" in s or "licenc" in s:
        return "Afastado"
    if "inativ" in s or "deslig" in s or "rescid" in s or "demit" in s:
        return "Inativo"
    return to_text(status_txt) or "Ativo"


def get_sheet(wb):
    # 1) tenta aba preferida
    if ABA_PREFERIDA and ABA_PREFERIDA in wb.sheetnames:
        return wb[ABA_PREFERIDA]

    # 2) tenta achar por cabeçalhos
    wanted_hits = ("nome", "funcao", "salario")
    for sh in wb.worksheets:
        for row in sh.iter_rows(min_row=1, max_row=20, values_only=True):
            headers = [norm_header(x) for x in row if x is not None and str(x).strip() != ""]
            if not headers:
                continue
            joined = " | ".join(headers)
            hits = sum(1 for w in wanted_hits if w in joined)
            if hits >= 2:
                return sh

    return wb.active


def detect_header_row(sheet):
    if LINHA_CABECALHO_FIXA:
        return LINHA_CABECALHO_FIXA

    # tenta achar uma linha que tenha "nome" e "salario"
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        vals = [norm_header(x) for x in row if x is not None and str(x).strip() != ""]
        joined = " ".join(vals)
        if ("nome" in joined) and ("salario" in joined or "salário" in joined):
            return i

    # fallback: primeira linha não vazia
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        if any(x not in (None, "", " ") for x in row):
            return i

    return 1


def build_header_map(sheet, header_row_idx):
    """header_normalizado -> col_idx (1-based)"""
    row = list(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
    hmap = {}
    for col_idx, h in enumerate(row, start=1):
        key = norm_header(h)
        if key:
            hmap[key] = col_idx
    return hmap


def col(hmap, *names):
    for n in names:
        nn = norm_header(n)
        if nn in hmap:
            return hmap[nn]
    return None


def safe_cell_value(values_sheet, formulas_sheet, r, c):
    """
    values_sheet: wb com data_only=True (valores)
    formulas_sheet: wb com data_only=False (para detectar fórmula)
    """
    v = values_sheet.cell(row=r, column=c).value
    if v is not None:
        return v

    f = formulas_sheet.cell(row=r, column=c).value
    if isinstance(f, str) and f.startswith("="):
        return None
    return None


def read_rows(values_sh, formulas_sh):
    header_row = detect_header_row(values_sh)
    hmap = build_header_map(values_sh, header_row)

    # tenta por cabeçalho
    c_setor  = col(hmap, "Setor")
    c_equipe = col(hmap, "Equipe")
    c_custo  = col(hmap, "Custo")
    c_matric = col(hmap, "Matricula", "Matrícula")
    c_nome   = col(hmap, "Nome")
    c_funcao = col(hmap, "Funcao", "Função")
    c_adm    = col(hmap, "Admissao", "Admissão")
    c_resc   = col(hmap, "Rescisao", "Rescisão")
    c_stat   = col(hmap, "Status", "Situação")
    c_hor    = col(hmap, "Horario", "Horário")
    c_reg    = col(hmap, "Regime", "Regime de Contratacao", "Regime de Contratação", "Vínculo", "Vinculo")
    c_sal    = col(hmap, "Salario", "Salário")
    c_vr     = col(hmap, "ValeRefeicao", "Vale Refeicao", "Vale Refeição", "VR")
    c_valv   = col(hmap, "ValorVeiculo", "Valor Veiculo", "Valor Veículo", "Frota")
    c_comb   = col(hmap, "Combustivel", "Combustível")
    c_veic   = col(hmap, "Veiculo", "Veículo")
    c_placa  = col(hmap, "Placa")
    c_valm   = col(hmap, "ValorMensal", "Valor Mensal", "valor_mensal")
    c_cfrota = col(hmap, "CustoFrota", "Custo Frota", "custo_frota")
    c_catf   = col(hmap, "CategoriaFrota", "Categoria Frota", "categoria_frota", "categoria")
    c_tipoeq = col(hmap, "TipoEquipamento", "Tipo Equipamento", "tipo_equipamento")

    # fallback por letra (se não achou pelo cabeçalho)
    def idx_or_fallback(current, key):
        if current is not None:
            return current
        letra = COL_FIXA.get(key)
        return column_index_from_string(letra) if letra else None

    c_setor  = idx_or_fallback(c_setor,  "Setor")
    c_equipe = idx_or_fallback(c_equipe, "Equipe")
    c_custo  = idx_or_fallback(c_custo,  "Custo")
    c_matric = idx_or_fallback(c_matric, "Matricula")
    c_nome   = idx_or_fallback(c_nome,   "Nome")
    c_funcao = idx_or_fallback(c_funcao, "Funcao")
    c_adm    = idx_or_fallback(c_adm,    "Admissao")
    c_resc   = idx_or_fallback(c_resc,   "Rescisao")
    c_stat   = idx_or_fallback(c_stat,   "Status")
    c_hor    = idx_or_fallback(c_hor,    "Horario")
    c_reg    = idx_or_fallback(c_reg,    "Regime")
    c_sal    = idx_or_fallback(c_sal,    "Salario")
    c_vr     = idx_or_fallback(c_vr,     "ValeRefeicao")
    c_valv   = idx_or_fallback(c_valv,   "ValorVeiculo")
    c_comb   = idx_or_fallback(c_comb,   "Combustivel")
    c_veic   = idx_or_fallback(c_veic,   "Veiculo")
    c_placa  = idx_or_fallback(c_placa,  "Placa")

    print(f"✅ Cabeçalho na linha: {header_row}")
    print(f"✅ Colunas (1-based): Nome={c_nome} | Salário={c_sal} | VR={c_vr} | Frota={c_valv} | Combustível={c_comb}")

    # detectar fórmulas sem resultado gravado (para avisar)
    formula_sem_cache = 0

    registros = []
    max_row = values_sh.max_row

    for r in range(header_row + 1, max_row + 1):
        nome_val = safe_cell_value(values_sh, formulas_sh, r, c_nome) if c_nome else None
        nome = to_text(nome_val)
        if nome == "":
            continue

        # aviso de fórmula sem cache (principalmente salário)
        if c_sal:
            sal_raw = safe_cell_value(values_sh, formulas_sh, r, c_sal)
            if sal_raw is None:
                f = formulas_sh.cell(row=r, column=c_sal).value
                if isinstance(f, str) and f.startswith("="):
                    formula_sem_cache += 1

        def get_text(c):
            if not c:
                return ""
            return to_text(safe_cell_value(values_sh, formulas_sh, r, c))

        def get_money(c):
            if not c:
                return 0.0
            v = safe_cell_value(values_sh, formulas_sh, r, c)
            return to_float(v)

        setor = get_text(c_setor)
        equipe = get_text(c_equipe)
        custo = get_text(c_custo)
        matric = get_text(c_matric)
        funcao = get_text(c_funcao)
        adm = get_text(c_adm)
        resc = get_text(c_resc)
        status = normalize_status(get_text(c_stat))
        regime = get_text(c_reg)
        tipo = detect_tipo(regime)

        categoria_frota = get_text(c_catf)
        tipo_equipamento = get_text(c_tipoeq)

        salario = get_money(c_sal)
        vr = get_money(c_vr)
        combustivel = get_money(c_comb)  # ✅ agora é número

        # ===================== FROTA / VEÍCULOS =====================
        # A planilha tem:
        # - custo_frota: custo mensal consolidado (inclui combustível na frota leve)
        # - valor_mensal: locação/contrato (ex.: caminhões e máquinas)
        # - valor_veiculo: custo fixo do veículo (quando aplicável)
        custo_frota = get_money(c_cfrota) if c_cfrota else 0.0
        valor_mensal = get_money(c_valm) if c_valm else 0.0
        valor_veiculo = get_money(c_valv) if c_valv else 0.0

        # Prioridade para manter coerência com Excel:
        # 1) custo_frota (quando existir)
        # 2) valor_mensal (quando existir e >0)
        # 3) valor_veiculo (fallback)
        if c_cfrota:
            frota = custo_frota
        elif (c_valm and valor_mensal > 0):
            frota = valor_mensal
        else:
            frota = valor_veiculo
        # ============================================================

        # Importante: combustível fica como indicador separado no JSON,
        # mas o custo mensal total deve somar SALÁRIO + VR + FROTA,
        # pois na frota leve o combustível já está incorporado em custo_frota.
        custo_total = salario + vr + frota

        registros.append({
            "setor": setor,
            "equipe": equipe,
            "custo": custo,
            "matricula": matric,
            "nome": nome,
            "funcao": funcao,
            "admissao": adm,
            "rescisao": resc,
            "status": status,   # Ativo / Afastado / Inativo
            "regime": regime,
            "tipo": tipo,       # CLT / PJ / ""
            "salario": round(salario, 2),
            "vr": round(vr, 2),
            "frota": round(frota, 2),
            "combustivel": round(combustivel, 2),
            "valor_mensal": round(valor_mensal, 2),
            "custo_frota": round(frota, 2),
            "categoria_frota": categoria_frota,
            "tipo_equipamento": tipo_equipamento,
            "custo_total": round(custo_total, 2),
            "horario": get_text(c_hor),
            "veiculo": get_text(c_veic),
            "placa": get_text(c_placa),
        })

    if formula_sem_cache > 0:
        print("⚠️ ATENÇÃO: encontrei células com FÓRMULA sem resultado gravado no arquivo.")
        print(f"➡️ Quantidade (aprox): {formula_sem_cache}")
        print("✅ SOLUÇÃO: abra o Excel > Arquivo > Salvar e feche. Depois rode o script de novo.")
        print("   (Isso grava o resultado calculado e o openpyxl consegue ler.)")

    return registros


def main():
    if ARQ_EXCEL is None:
        raise FileNotFoundError(
            "Não encontrei o Excel de funcionários.\n"
            "Procurei estes caminhos:\n- " + "\n- ".join(str(p) for p in CANDIDATOS_EXCEL)
        )

    # 1) arquivo com valores
    wb_values = load_workbook(ARQ_EXCEL, data_only=True)
    # 2) arquivo com fórmulas (para detectar se está vindo None por fórmula)
    wb_formulas = load_workbook(ARQ_EXCEL, data_only=False)

    sh_values = get_sheet(wb_values)
    sh_formulas = wb_formulas[sh_values.title]

    print("✅ Excel lido de:", ARQ_EXCEL)
    print("✅ Aba detectada:", sh_values.title)

    registros = read_rows(sh_values, sh_formulas)

    mes = infer_mes_competencia()

    # somar somente ativos + afastados (inativo não entra em custo mensal)
    def conta_custo(r):
        s = norm_header(r.get("status", ""))
        return s in ("ativo", "afastado")

    ativos = sum(1 for r in registros if norm_header(r.get("status")) == "ativo")
    afastados = sum(1 for r in registros if norm_header(r.get("status")) == "afastado")

    total_salarios = round(sum(r["salario"] for r in registros if conta_custo(r)), 2)
    total_vr = round(sum(r["vr"] for r in registros if conta_custo(r)), 2)
    total_combustivel = round(sum(r["combustivel"] for r in registros if conta_custo(r)), 2)
    total_frota = round(sum(r["frota"] for r in registros if conta_custo(r)), 2)
    custo_mensal_total = round(total_salarios + total_vr + total_frota, 2)  # combustível já está na frota leve

    salario_clt = round(sum(r["salario"] for r in registros if conta_custo(r) and r.get("tipo") == "CLT"), 2)
    salario_pj = round(sum(r["salario"] for r in registros if conta_custo(r) and r.get("tipo") == "PJ"), 2)

    # agregações úteis p/ dashboard
    por_setor = {}
    por_equipe = {}
    por_tipo = {"CLT": 0.0, "PJ": 0.0, "OUTROS": 0.0}

    for r in registros:
        if not conta_custo(r):
            continue

        setor = r.get("setor") or "Sem Setor"
        equipe = r.get("equipe") or "Sem Equipe"
        tipo = r.get("tipo") or "OUTROS"
        if tipo not in ("CLT", "PJ"):
            tipo = "OUTROS"

        por_setor[setor] = por_setor.get(setor, 0.0) + r["custo_total"]
        por_equipe[equipe] = por_equipe.get(equipe, 0.0) + r["custo_total"]
        por_tipo[tipo] = por_tipo.get(tipo, 0.0) + r["custo_total"]

    # funcionarios.json (principal)
    out = {
        "atualizado_em": now_iso_local(),
        "mes": mes,
        "resumo": {
            "ativos": ativos,
            "afastados": afastados,
            "total_salarios": total_salarios,
            "total_vr": total_vr,
            "total_combustivel": total_combustivel,
            "total_frota": total_frota,
            "custo_mensal_total": custo_mensal_total,
            "salario_clt": salario_clt,
            "salario_pj": salario_pj,
            "qtd_registros": len(registros),
        },
        "por_setor": {k: round(v, 2) for k, v in sorted(por_setor.items())},
        "por_equipe": {k: round(v, 2) for k, v in sorted(por_equipe.items())},
        "por_tipo": {k: round(v, 2) for k, v in por_tipo.items()},
        "funcionarios": registros,
    }

    ARQUIVO_SAIDA.parent.mkdir(parents=True, exist_ok=True)
    with open(ARQUIVO_SAIDA, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    # funcionarios_historico.json (append/merge por mês)
    historico = {"atualizado_em": now_iso_utc(), "series": []}

    if ARQUIVO_HIST.exists():
        try:
            with open(ARQUIVO_HIST, "r", encoding="utf-8") as f:
                historico = json.load(f) or historico
            if "series" not in historico or not isinstance(historico["series"], list):
                historico["series"] = []
        except:
            historico = {"atualizado_em": now_iso_utc(), "series": []}

    item_mes = {
        "mes": mes,
        "atualizado_em": now_iso_local(),
        "ativos": ativos,
        "afastados": afastados,
        "total_salarios": total_salarios,
        "total_vr": total_vr,
        "total_combustivel": total_combustivel,
        "total_frota": total_frota,
        "custo_mensal_total": custo_mensal_total,
        "salario_clt": salario_clt,
        "salario_pj": salario_pj,
    }

    replaced = False
    for i, it in enumerate(historico["series"]):
        if isinstance(it, dict) and it.get("mes") == mes:
            historico["series"][i] = item_mes
            replaced = True
            break
    if not replaced:
        historico["series"].append(item_mes)

    historico["series"] = sorted(historico["series"], key=lambda x: x.get("mes", ""))
    historico["atualizado_em"] = now_iso_utc()

    with open(ARQUIVO_HIST, "w", encoding="utf-8") as f:
        json.dump(historico, f, ensure_ascii=False, indent=2)

    print("✅ Gerado:", ARQUIVO_SAIDA)
    print("✅ Histórico atualizado:", ARQUIVO_HIST)
    print("➡️ Resumo:", out["resumo"])


if __name__ == "__main__":
    main()