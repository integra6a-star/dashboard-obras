import json
from pathlib import Path
import openpyxl
from datetime import datetime, timezone, timedelta

BASE = Path(__file__).resolve().parent
ARQUIVO_EXCEL = BASE / "EAP_PRODUCAO.xlsx"
ARQUIVO_JSON = BASE / "dados.json"

MESES_ORDEM = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]


def carregar_curva_excel(caminho_excel: Path):
    wb = openpyxl.load_workbook(caminho_excel, data_only=True)
    ws = wb[wb.sheetnames[0]]

    cabecalho_linha = None
    indices = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        valores = [str(v).strip().lower() if v is not None else "" for v in row]
        if "ano" in valores and "mes" in valores and "eap" in valores and "produzido" in valores:
            cabecalho_linha = row_idx
            indices = {
                "ano": valores.index("ano"),
                "mes": valores.index("mes"),
                "eap": valores.index("eap"),
                "produzido": valores.index("produzido"),
            }
            break

    if cabecalho_linha is None:
        raise ValueError("Não encontrei as colunas Ano, Mes, EAP e Produzido na planilha.")

    mensal = []
    for row in ws.iter_rows(min_row=cabecalho_linha + 1, values_only=True):
        ano = row[indices["ano"]] if indices["ano"] < len(row) else None
        mes = row[indices["mes"]] if indices["mes"] < len(row) else None
        eap = row[indices["eap"]] if indices["eap"] < len(row) else None
        produzido = row[indices["produzido"]] if indices["produzido"] < len(row) else None

        if ano in (None, "") and mes in (None, ""):
            continue

        mensal.append({
            "ano": int(ano),
            "mes": str(mes).strip(),
            "eap": float(eap or 0),
            "produzido": float(produzido or 0),
        })

    mensal.sort(key=lambda x: (x["ano"], MESES_ORDEM.index(x["mes"]) if x["mes"] in MESES_ORDEM else 99))
    return mensal


def atualizar_json():
    if not ARQUIVO_EXCEL.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {ARQUIVO_EXCEL}")
    if not ARQUIVO_JSON.exists():
        raise FileNotFoundError(f"JSON não encontrado: {ARQUIVO_JSON}")

    mensal = carregar_curva_excel(ARQUIVO_EXCEL)

    with ARQUIVO_JSON.open("r", encoding="utf-8") as f:
        dados = json.load(f)

    dados["eap_producao"] = {"mensal": mensal}

    tz = timezone(timedelta(hours=-3))
    dados["atualizado_em"] = datetime.now(tz).isoformat()

    with ARQUIVO_JSON.open("w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

    print("dados.json atualizado com sucesso.")
    print(f"Meses carregados: {len(mensal)}")


if __name__ == "__main__":
    atualizar_json()
