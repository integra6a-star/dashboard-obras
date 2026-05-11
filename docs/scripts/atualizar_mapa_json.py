import json
import shutil
from pathlib import Path
from openpyxl import load_workbook

ARQUIVO_EXCEL = "planilha_base_mapa.xlsx"
ARQUIVO_JSON = "dados_mapa.json"

def ler_aba(ws):
    cabecalho = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    registros = []
    for r in range(2, ws.max_row + 1):
        item = {}
        vazio = True
        for c, nome in enumerate(cabecalho, start=1):
            if not nome:
                continue
            valor = ws.cell(r, c).value
            if valor is not None:
                vazio = False
            item[nome] = valor
        if not vazio:
            registros.append(item)
    return registros

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent
DOCS_DIR = ROOT_DIR / "docs"

def escolher_planilha(nome):
    candidatos = [ROOT_DIR / nome, DOCS_DIR / nome]
    existentes = [p for p in candidatos if p.exists()]
    if not existentes:
        raise FileNotFoundError(f"Planilha nao encontrada na raiz nem em docs: {nome}")
    origem = max(existentes, key=lambda p: p.stat().st_mtime)
    for destino in candidatos:
        if destino == origem:
            continue
        if not destino.exists() or origem.stat().st_mtime > destino.stat().st_mtime + 1:
            destino.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(origem, destino)
            print(f"Sincronizado {origem.name}: {origem.parent.name} -> {destino.parent.name}")
    return origem


excel = escolher_planilha(ARQUIVO_EXCEL)
wb = load_workbook(excel, data_only=True)

dados = {
    "obras": ler_aba(wb["OBRAS"]),
    "pontos": ler_aba(wb["PONTOS"]),
    "trechos": ler_aba(wb["TRECHOS"]),
}

for base in (ROOT_DIR, DOCS_DIR):
    base.mkdir(parents=True, exist_ok=True)
    with open(base / ARQUIVO_JSON, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

print(f"JSON atualizado com sucesso: {ARQUIVO_JSON}")
