import json
from pathlib import Path
from datetime import datetime
import openpyxl

# Ajuste o nome do arquivo se necessário
ARQUIVO_PDS = "pds_word_completo_preenchido(2)(1).xlsx"
ABA = "PDS_Lancamentos"
SAIDA = "pds.json"

def norm(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()

def mes_ref(data_txt):
    try:
        return datetime.strptime(data_txt[:10], "%Y-%m-%d").strftime("%Y-%m")
    except Exception:
        return ""

def main():
    caminho = Path(ARQUIVO_PDS)
    if not caminho.exists():
        raise FileNotFoundError(f"Não encontrei {ARQUIVO_PDS} na pasta.")

    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[ABA]

    headers = [norm(ws.cell(4, c).value) for c in range(1, ws.max_column + 1)]
    dados = []

    for row in ws.iter_rows(min_row=5, values_only=True):
        item = {}
        vazio = True
        for h, v in zip(headers, row):
            if not h:
                continue
            valor = norm(v)
            if valor:
                vazio = False
            item[h] = valor

        if vazio:
            continue

        item["Mes_Ref"] = mes_ref(item.get("Data", ""))
        # nomes amigáveis para o relatório
        item["Obra"] = item.get("Obra", "")
        item["Atividade"] = item.get("Atividade", "")
        item["Tipo_Atividade"] = item.get("Tipo_Atividade", "")
        item["PV_Inicio"] = item.get("PV_Inicio", "")
        item["PV_Fim"] = item.get("PV_Fim", "")
        item["PV_Texto"] = item.get("PV_Texto", "")
        item["Rua"] = item.get("RUA", item.get("Rua", ""))
        item["Numero"] = item.get("NUMERO", item.get("Numero", ""))
        item["Municipio"] = item.get("MUNICIPIO", item.get("Municipio", ""))
        item["Subprefeitura"] = item.get("SUBPREFEITURA", item.get("Subprefeitura", ""))
        dados.append(item)

    Path(SAIDA).write_text(json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Gerado {SAIDA} com {len(dados)} registros.")

if __name__ == "__main__":
    main()
