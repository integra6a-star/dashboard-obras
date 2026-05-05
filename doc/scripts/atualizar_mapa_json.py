import json
from pathlib import Path
from openpyxl import load_workbook

ARQUIVO_EXCEL = "planilha_base_mapa.xlsx"
ARQUIVO_JSON = "dados_mapa.json"

def ler_aba(ws):
    cabecalho = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    registros = []
    for r in range(2, ws.max_row + 1):
        item = {}
        vazio = True
        for c, nome in enumerate(cabecalho, start=1):
            if not nome:
                continue
            valor = ws.cell(r, c).value
            if valor is not None:
                vazio = False
            item[nome] = valor
        if not vazio:
            registros.append(item)
    return registros

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent
DOCS_DIR = ROOT_DIR / "docs"

base = DOCS_DIR if (DOCS_DIR / ARQUIVO_EXCEL).exists() else ROOT_DIR
wb = load_workbook(base / ARQUIVO_EXCEL, data_only=True)

dados = {
    "obras": ler_aba(wb["OBRAS"]),
    "pontos": ler_aba(wb["PONTOS"]),
    "trechos": ler_aba(wb["TRECHOS"]),
}

with open(base / ARQUIVO_JSON, "w", encoding="utf-8") as f:
    json.dump(dados, f, ensure_ascii=False, indent=2)

print(f"JSON atualizado com sucesso: {ARQUIVO_JSON}")
