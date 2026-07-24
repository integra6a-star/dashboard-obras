import io
import json
import math
import shutil
import struct
import zipfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
ZIP_SHAPE = Path(r"C:\Users\micro\Downloads\SHAPE_RCE CJ.UNIÃO.zip")
PLANILHA = ROOT / "planilha_base_mapa.xlsx"
OBRA_ID = "conjunto_uniao"
NOME_OBRA = "RCE Conjunto União"


def to_float(value):
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def to_number(value):
    numero = to_float(value)
    if numero is None:
        return None
    return int(numero) if abs(numero - int(numero)) < 1e-9 else numero


def utm23s_to_latlon(easting, northing):
    # SIRGAS 2000 / UTM 23S is practically WGS84 for dashboard precision.
    a = 6378137.0
    f = 1 / 298.257222101
    k0 = 0.9996
    e = math.sqrt(f * (2 - f))
    e1 = (1 - math.sqrt(1 - e * e)) / (1 + math.sqrt(1 - e * e))
    x = easting - 500000.0
    y = northing - 10000000.0
    lon0 = math.radians(-45)
    m = y / k0
    mu = m / (a * (1 - e**2 / 4 - 3 * e**4 / 64 - 5 * e**6 / 256))
    fp = (
        mu
        + (3 * e1 / 2 - 27 * e1**3 / 32) * math.sin(2 * mu)
        + (21 * e1**2 / 16 - 55 * e1**4 / 32) * math.sin(4 * mu)
        + (151 * e1**3 / 96) * math.sin(6 * mu)
        + (1097 * e1**4 / 512) * math.sin(8 * mu)
    )
    ep2 = e * e / (1 - e * e)
    c1 = ep2 * math.cos(fp) ** 2
    t1 = math.tan(fp) ** 2
    r1 = a * (1 - e * e) / ((1 - e * e * math.sin(fp) ** 2) ** 1.5)
    n1 = a / math.sqrt(1 - e * e * math.sin(fp) ** 2)
    d = x / (n1 * k0)
    lat = fp - (n1 * math.tan(fp) / r1) * (
        d**2 / 2
        - (5 + 3 * t1 + 10 * c1 - 4 * c1**2 - 9 * ep2) * d**4 / 24
        + (61 + 90 * t1 + 298 * c1 + 45 * t1**2 - 252 * ep2 - 3 * c1**2) * d**6 / 720
    )
    lon = lon0 + (
        d
        - (1 + 2 * t1 + c1) * d**3 / 6
        + (5 - 2 * c1 + 28 * t1 - 3 * c1**2 + 8 * ep2 + 24 * t1**2) * d**5 / 120
    ) / math.cos(fp)
    return round(math.degrees(lat), 7), round(math.degrees(lon), 7)


def read_dbf(data, encoding="iso-8859-1"):
    total = struct.unpack("<I", data[4:8])[0]
    header_len = struct.unpack("<H", data[8:10])[0]
    row_len = struct.unpack("<H", data[10:12])[0]
    fields = []
    offset = 1
    pos = 32
    while data[pos] != 0x0D:
        name = data[pos : pos + 11].split(b"\0", 1)[0].decode(encoding, "replace")
        kind = chr(data[pos + 11])
        length = data[pos + 16]
        fields.append((name, kind, length, offset))
        offset += length
        pos += 32

    rows = []
    for idx in range(total):
        raw_row = data[header_len + idx * row_len : header_len + (idx + 1) * row_len]
        if not raw_row or raw_row[0:1] == b"*":
            continue
        row = {}
        for name, kind, length, offset in fields:
            raw = raw_row[offset : offset + length].decode(encoding, "replace").strip()
            row[name] = raw
        rows.append(row)
    return rows


def read_shp(data):
    records = []
    pos = 100
    while pos + 8 <= len(data):
        _, content_len = struct.unpack(">2i", data[pos : pos + 8])
        pos += 8
        content = data[pos : pos + content_len * 2]
        pos += content_len * 2
        shape_type = struct.unpack("<i", content[:4])[0]
        if shape_type == 1:
            records.append([struct.unpack("<2d", content[4:20])])
        elif shape_type in (3, 5):
            parts_count, points_count = struct.unpack("<2i", content[36:44])
            points_offset = 44 + 4 * parts_count
            records.append([
                struct.unpack("<2d", content[points_offset + i * 16 : points_offset + i * 16 + 16])
                for i in range(points_count)
            ])
        else:
            records.append([])
    return records


def nested_layers():
    with zipfile.ZipFile(ZIP_SHAPE) as outer:
        for outer_name in outer.namelist():
            with zipfile.ZipFile(io.BytesIO(outer.read(outer_name))) as inner:
                names = inner.namelist()
                shp = next(name for name in names if name.lower().endswith(".shp"))
                dbf = next(name for name in names if name.lower().endswith(".dbf"))
                cpg = next((name for name in names if name.lower().endswith(".cpg")), None)
                encoding = inner.read(cpg).decode("ascii", "ignore").strip() if cpg else "iso-8859-1"
                if encoding.upper() == "ISO-8859-1":
                    encoding = "iso-8859-1"
                yield outer_name, read_dbf(inner.read(dbf), encoding), read_shp(inner.read(shp))


def ponto_nome(row):
    observ = (row.get("Observ") or "").strip()
    if observ:
        return observ.replace(" ", "")
    if row.get("N_PV"):
        return f"PV-{int(float(row['N_PV']))}"
    if row.get("N_PI"):
        return f"PI-{int(float(row['N_PI']))}"
    return ""


def ponto_tipo(nome):
    return "PV" if nome.upper().startswith("PV") else "PI"


def coord_key(easting, northing):
    return (round(easting, 3), round(northing, 3))


def get_headers(ws):
    return [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]


def rewrite_sheet(ws, rows):
    headers = get_headers(ws)
    ws.delete_rows(2, ws.max_row)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def load_sheet_rows(ws):
    headers = get_headers(ws)
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values):
            continue
        rows.append({header: values[idx] if idx < len(values) else None for idx, header in enumerate(headers)})
    return rows


def main():
    if not ZIP_SHAPE.exists():
        raise FileNotFoundError(ZIP_SHAPE)

    pontos_by_name = {}
    pontos_by_coord = {}
    linhas = []

    for layer_name, rows, geometries in nested_layers():
        is_rede = "REDE" in layer_name.upper()
        for idx, (row, geom) in enumerate(zip(rows, geometries), start=1):
            if is_rede:
                if len(geom) < 2:
                    continue
                linhas.append((row, geom))
                continue

            nome = ponto_nome(row)
            if not nome:
                continue
            easting = to_float(row.get("Coord E")) or (geom[0][0] if geom else None)
            northing = to_float(row.get("Coord N")) or (geom[0][1] if geom else None)
            if easting is None or northing is None:
                continue
            if not (300000 <= easting <= 450000 and 7300000 <= northing <= 7500000):
                continue
            if nome in pontos_by_name:
                continue

            lat, lon = utm23s_to_latlon(easting, northing)
            ponto = {
                "obra_id": OBRA_ID,
                "tipo_ponto": ponto_tipo(nome),
                "ponto_id": nome,
                "nome_ponto": nome,
                "coord_n": round(northing, 3),
                "coord_e": round(easting, 3),
                "latitude": lat,
                "longitude": lon,
                "cota_tampa": to_float(row.get("Cota_Tampa")),
                "cota_fundo": to_float(row.get("Cota_fundo")),
                "profundidade_m": to_float(row.get("Prof")),
                "observacao": nome,
            }
            pontos_by_name[nome] = ponto
            pontos_by_coord[coord_key(easting, northing)] = ponto

    trechos = []
    for idx, (row, geom) in enumerate(linhas, start=1):
        inicio_xy = geom[0]
        fim_xy = geom[-1]
        inicio = pontos_by_coord.get(coord_key(*inicio_xy))
        fim = pontos_by_coord.get(coord_key(*fim_xy))
        lat_inicio, lon_inicio = utm23s_to_latlon(*inicio_xy)
        lat_fim, lon_fim = utm23s_to_latlon(*fim_xy)
        pv_inicio = inicio["nome_ponto"] if inicio else ""
        pv_fim = fim["nome_ponto"] if fim else ""
        trechos.append({
            "obra_id": OBRA_ID,
            "trecho_id": f"CU-{idx:03d}",
            "trecho_nome": f"{pv_inicio} - {pv_fim}".strip(" -"),
            "pv_inicio": pv_inicio,
            "pv_fim": pv_fim,
            "material": row.get("Material") or "",
            "dn": to_number(row.get("Diametro")),
            "metodo": row.get("Metod_Cons") or "",
            "extensao_m": to_float(row.get("Comp_Real")),
            "prof_inicial_m": to_float(row.get("Prof_GI_M")),
            "prof_final_m": to_float(row.get("Prof_GI_J")),
            "prof_max_m": None,
            "status": "concluída" if "Opera" in (row.get("Estado") or "") else row.get("Estado"),
            "observacao": row.get("Observ") or None,
            "lat_inicio": lat_inicio,
            "lon_inicio": lon_inicio,
            "lat_fim": lat_fim,
            "lon_fim": lon_fim,
        })

    wb = load_workbook(PLANILHA)
    obras = [row for row in load_sheet_rows(wb["OBRAS"]) if row.get("obra_id") != OBRA_ID]
    pontos = [row for row in load_sheet_rows(wb["PONTOS"]) if row.get("obra_id") != OBRA_ID]
    trechos_existentes = [row for row in load_sheet_rows(wb["TRECHOS"]) if row.get("obra_id") != OBRA_ID]

    obras.append({
        "obra_id": OBRA_ID,
        "nome_obra": NOME_OBRA,
        "tipo": "RCE",
        "subtitulo": "Base real do shape: SIRGAS 2000 / UTM zona 23S convertida para latitude/longitude.",
        "ativa": "SIM",
    })
    pontos.extend(pontos_by_name[name] for name in sorted(pontos_by_name))
    trechos_existentes.extend(trechos)

    rewrite_sheet(wb["OBRAS"], obras)
    rewrite_sheet(wb["PONTOS"], pontos)
    rewrite_sheet(wb["TRECHOS"], trechos_existentes)
    wb.save(PLANILHA)

    DOCS.mkdir(exist_ok=True)
    shutil.copy2(PLANILHA, DOCS / PLANILHA.name)

    payload = {
        "obras": obras,
        "pontos": pontos,
        "trechos": trechos_existentes,
    }
    for base in (ROOT, DOCS):
        with (base / "dados_mapa.json").open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"{NOME_OBRA}: {len(pontos_by_name)} ponto(s), {len(trechos)} trecho(s).")
    print(f"dados_mapa.json: {len(obras)} obra(s), {len(pontos)} ponto(s), {len(trechos_existentes)} trecho(s).")


if __name__ == "__main__":
    main()
