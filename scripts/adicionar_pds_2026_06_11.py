# -*- coding: utf-8 -*-
"""Adiciona a PDS informada por chat em 11/06/2026."""

from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
WORKBOOK = DOCS / "pds_word_completo_preenchido(2).xlsx"
ROOT_COPY = ROOT / WORKBOOK.name
SHEET_NAME = "PDS_Lancamentos"
SOURCE = "PDS-Chat-2026-06-11"
DATA = datetime(2026, 6, 11)

PV_RANGE_RE = re.compile(
    r"\b(?:PV|PVE|PI|PVI|PVS)[-\s]*([A-Z]?\d+(?:[.,]\d+)*)\s*(?:ao|a|-|e)\s*(?:PV|PVE|PI|PVI|PVS)?[-\s]*([A-Z]?\d+(?:[.,]\d+)*)",
    re.I,
)
PV_RE = re.compile(r"\b(?:PV|PVE|PI|PVI|PVS)[-\s]*([A-Z]?\d+(?:[.,]\d+)*)\b", re.I)


ROWS = [
    ("CT João Canzi", "Robson", "Márcio", "Corte da cabeça de puxe PV-19", "Planejamento"),
    ("CT João Canzi", "Robson", "Medeiros", "Reaterro Shaft PV 10.1", "Planejamento"),
    ("CT João Canzi", "Robson", "Cidão", "Transformação PV 18", "Planejamento"),
    ("CT João Canzi", "Robson", "Márcio", "Shaft PV 21", "Planejamento"),
    ("RCE Elvira", "Silvestre", "Valter", "VCA, PVs e Ligações", "Planejamento"),
    ("RCE Elias Sleiman / CT João Canzi", "Silvestre", "Edvando", "Interligação PV 17", "Planejamento"),
    ("RCE Conj. União", "Caroline", "Jhonatan", "Shaft PV 48", "Planejamento"),
    ("RCE Américo Machado/Elis Regina", "Caroline", "Cesar", "Transformação PV 01", "Planejamento"),
    ("CTS Lourdes Complementar", "Robson", "Marcelo", "Interligação entre PV-05 e PV-E", "Planejamento"),
    ("RCE São Lucas", "Fábio", "Ricardo (PI 10 ao PI 11)", "VCA, PVs e Ligações", "Planejamento"),
    ("RCE São Lucas", "Fábio", "Claudinei (PI 38 ao PI 81)", "VCA, PVs e Ligações", "Planejamento"),
    ("RCE São Lucas", "Fábio", "Miro (PI 61 ao PI 62)", "VCA, PVs e Ligações", "Planejamento"),
    ("RCE São Lucas", "Fábio", "Emanuel (PI 25 ao PI 26)", "VCA, PVs e Ligações", "Planejamento"),
    ("RCE São Lucas", "Fábio", "Antonio (PI 55 ao PI 56)", "VCA, PVs e Ligações", "Planejamento"),
    ("Serviços Reposição", "Leandro", "Serviços Reposição", "Reparos Gerais", "Planejamento"),
    ("Vacal", "", "Vacal", "João Canzi - Henko", "Apoio"),
    ("Vacal", "", "Vacal", "João Canzi - Henko", "Apoio"),
    ("Vacal", "", "Vacal", "ITI 15 - Henko (Osmar)", "Apoio"),
    ("Guindauto", "", "Guindauto", "Marcio - Apoio as equipes", "Apoio"),
    ("Guindauto", "", "Guindauto", "Luiz - Apoio as equipes", "Apoio"),
]


def extract_pvs(text: str) -> tuple[str, str, str]:
    match = PV_RANGE_RE.search(text)
    if match:
        start, end = match.group(1).replace(",", "."), match.group(2).replace(",", ".")
        return start, end, f"{start} ao {end}"
    matches = [m.group(1).replace(",", ".") for m in PV_RE.finditer(text)]
    if not matches:
        return "", "", ""
    if len(matches) == 1:
        return matches[0], "", matches[0]
    return matches[0], matches[-1], " ao ".join([matches[0], matches[-1]])


def classify_activity(text: str) -> str:
    upper = text.upper()
    if "SHAFT" in upper:
        return "Shaft"
    if "TRANSFORM" in upper:
        return "Transformação"
    if "FURO" in upper or "PUXE" in upper or "HDD" in upper:
        return "HDD"
    if "VCA" in upper:
        return "VCA"
    if "LIGA" in upper or "INTERLIGA" in upper:
        return "Ligação"
    if "REATERRO" in upper:
        return "Reaterro"
    return "Atividade"


def main() -> None:
    wb = load_workbook(WORKBOOK)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = {str(ws.cell(4, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(4, c).value}

    data_col = headers["Data"]
    fonte_col = headers["Fonte_Arquivo"]

    removed = 0
    for row_idx in range(ws.max_row, 4, -1):
        value = ws.cell(row_idx, data_col).value
        data_key = value.strftime("%Y-%m-%d") if hasattr(value, "strftime") else str(value)
        if data_key.startswith("2026-06-11") and ws.cell(row_idx, fonte_col).value == SOURCE:
            ws.delete_rows(row_idx, 1)
            removed += 1

    inserted = 0
    for obra, responsavel, equipe, atividade, categoria in ROWS:
        target = ws.max_row + 1
        pv_inicio, pv_fim, pv_texto = extract_pvs(f"{equipe} {atividade}")
        values = {
            "Data": DATA,
            "Dia_Semana": "Quinta",
            "Obra": obra,
            "Responsavel_Obra": responsavel,
            "Equipe": equipe,
            "Atividade": atividade,
            "PV_Inicio": pv_inicio,
            "PV_Fim": pv_fim,
            "PV_Texto": pv_texto,
            "Tipo_Atividade": classify_activity(atividade),
            "Categoria": categoria,
            "Fonte_Arquivo": SOURCE,
            "MES_REF": f'=IF($A{target}="","",TEXT($A{target},"mmmm/yyyy"))',
            "MUNICIPIO": f'=IFERROR(XLOOKUP($N{target},Base_Enderecos!$A:$A,Base_Enderecos!$B:$B,""),"")',
            "SUBPREFEITURA": f'=IFERROR(XLOOKUP($N{target},Base_Enderecos!$A:$A,Base_Enderecos!$C:$C,""),"")',
            "METODO_PADRAO": f'=IF(ISNUMBER(SEARCH("HDD",$F{target})),"HDD",IF(ISNUMBER(SEARCH("VCA",$F{target})),"VCA","MND"))',
            "CHAVE_RELATORIO": f'=TEXT($A{target},"yyyy-mm")&"|"&$C{target}&"|"&$I{target}',
        }
        for key, value in values.items():
            if key in headers:
                ws.cell(target, headers[key]).value = value
        inserted += 1

    wb.save(WORKBOOK)
    shutil.copy2(WORKBOOK, ROOT_COPY)
    print(f"PDS 2026-06-11 adicionada: inseridas={inserted}; removidas={removed}")


if __name__ == "__main__":
    main()
