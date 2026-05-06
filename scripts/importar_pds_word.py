# -*- coding: utf-8 -*-
"""Importa automaticamente a PDS em Word para a planilha consolidada.

Por padrão o script procura arquivos .docx na pasta Desktop/PDS e
importa somente a data mais recente encontrada em cada arquivo. Isso evita
duplicar meses antigos quando a PDS mensal é atualizada todos os dias.
"""

from __future__ import annotations

import argparse
import re
import shutil
import unicodedata
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
WORKBOOK = DOCS / "pds_word_completo_preenchido(2).xlsx"
DEFAULT_DOCX_DIR = Path.home() / "Desktop" / "PDS"
SHEET_NAME = "PDS_Lancamentos"

DATE_RE = re.compile(r"(\d{2}/\d{2}/\d{4})")
PV_RE = re.compile(r"\b(?:PV|PVE|PI|PVI|PVS)[-\s]*([A-Z]?\d+(?:[.,]\d+)*)\b", re.I)
PV_RANGE_RE = re.compile(
    r"\b(?:PV|PVE|PI|PVI|PVS)[-\s]*([A-Z]?\d+(?:[.,]\d+)*)\s*(?:ao|a|-)\s*(?:PV|PVE|PI|PVI|PVS)?[-\s]*([A-Z]?\d+(?:[.,]\d+)*)",
    re.I,
)


def strip_accents(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    return "".join(ch for ch in value if not unicodedata.combining(ch))


def clean_text(value: str) -> str:
    value = (value or "").replace("\xa0", " ").replace("■", " ")
    value = re.sub(r"\s+", " ", value).strip(" ;.-\t")
    return value


def title_name(value: str) -> str:
    raw = clean_text(value)
    if not raw:
        return ""
    keep_upper = {"CT", "CTS", "RCE", "PV", "PVE", "HDD", "VCA", "MD", "ME"}
    words = []
    for word in raw.split():
        token = word.strip()
        bare = strip_accents(token).upper().strip(".,;:/")
        if bare in keep_upper:
            words.append(bare)
        else:
            words.append(token[:1].upper() + token[1:].lower())
    return " ".join(words)


def is_section(line: str) -> bool:
    if not line or ":" in line:
        return False
    upper = strip_accents(line).upper()
    if upper in {"INTEGRA 6A", "PDS - PROGRAMACAO DIARIA DE SERVICOS"}:
        return False
    return upper.startswith(("CT ", "CTS ", "RCE ", "INTERLIG", "VACAL", "GUINDAUTO"))


def split_section(line: str) -> tuple[str, str]:
    line = clean_text(line)
    if "–" in line:
        obra, resp = line.split("–", 1)
    elif " - " in line:
        obra, resp = line.split(" - ", 1)
    else:
        obra, resp = line, ""
    return title_name(obra), title_name(resp)


def classify_activity(text: str) -> str:
    upper = strip_accents(text).upper()
    if "SHAFT" in upper:
        return "Shaft"
    if "TRANSFORM" in upper:
        return "Transformação"
    if "ACAB" in upper:
        return "Acabamento"
    if "LAVAGEM" in upper:
        return "Lavagem"
    if "FURO" in upper or "PUXE" in upper or "HDD" in upper:
        return "HDD"
    if "VCA" in upper:
        return "VCA"
    if "TUBO" in upper:
        return "Tubo"
    if "LIGAC" in upper:
        return "Ligação"
    if "ESCAV" in upper:
        return "Escavação"
    return "Atividade"


def extract_pvs(text: str) -> tuple[str, str, str]:
    match = PV_RANGE_RE.search(text)
    if match:
        start, end = match.group(1).replace(",", "."), match.group(2).replace(",", ".")
        return start, end, f"{start} ao {end}"
    matches = [m.group(1).replace(",", ".") for m in PV_RE.finditer(text)]
    if not matches:
        return "", "", ""
    if len(matches) == 1:
        return matches[0], "", matches[0]
    return matches[0], matches[-1], " ao ".join([matches[0], matches[-1]])


def weekday_pt(dt: datetime) -> str:
    return ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"][dt.weekday()]


def parse_docx(path: Path) -> list[dict]:
    lines = read_docx_lines(path)
    rows: list[dict] = []
    current_date: datetime | None = None
    current_obra = ""
    current_resp = ""
    current_team = ""

    for line in lines:
        if not line:
            continue
        date_match = DATE_RE.search(line)
        if date_match:
            current_date = datetime.strptime(date_match.group(1), "%d/%m/%Y")
            current_obra = ""
            current_resp = ""
            current_team = ""
            continue

        if current_date is None:
            continue

        upper = strip_accents(line).upper()
        if upper.startswith("EQUIPE"):
            current_team = title_name(re.sub(r"^Equipe:?", "", line, flags=re.I))
            continue

        if upper.startswith("VACAL:"):
            obra = title_name(line.split(":", 1)[1].split("–", 1)[0].split("-", 1)[0])
            rows.append(make_row(current_date, obra or "Vacal", "", "Vacal", line.split(":", 1)[1], "Apoio", path.name))
            continue

        if upper.startswith("GUINDAUTO:"):
            rows.append(make_row(current_date, "Guindauto", "", "Guindauto", line.split(":", 1)[1], "Apoio", path.name))
            continue

        if is_section(line):
            current_obra, current_resp = split_section(line)
            current_team = ""
            continue

        if current_obra and (line.startswith("■") or current_team):
            rows.append(make_row(current_date, current_obra, current_resp, current_team, line, "Planejamento", path.name))

    return rows


def read_docx_lines(path: Path) -> list[str]:
    """Extrai parágrafos do .docx usando apenas bibliotecas padrão."""
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    lines: list[str] = []
    for paragraph in root.findall(".//w:p", ns):
        parts = []
        for node in paragraph.findall(".//w:t", ns):
            parts.append(node.text or "")
        text = clean_text("".join(parts))
        if text:
            lines.append(text)
    return lines


def make_row(dt: datetime, obra: str, resp: str, team: str, activity: str, category: str, source: str) -> dict:
    pv_start, pv_end, pv_text = extract_pvs(activity)
    return {
        "Data": dt,
        "Dia_Semana": weekday_pt(dt),
        "Obra": obra,
        "Responsavel_Obra": resp,
        "Equipe": team,
        "Atividade": clean_text(activity),
        "PV_Inicio": pv_start,
        "PV_Fim": pv_end,
        "PV_Texto": pv_text or pv_start or pv_end,
        "Tipo_Atividade": classify_activity(activity),
        "Categoria": category,
        "Fonte_Arquivo": source,
    }


def load_headers(ws) -> dict[str, int]:
    return {str(ws.cell(4, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(4, c).value}


def row_key(row: dict) -> tuple:
    dt = row["Data"].strftime("%Y-%m-%d") if hasattr(row["Data"], "strftime") else str(row["Data"])
    return (
        dt,
        row.get("Obra", ""),
        row.get("Equipe", ""),
        row.get("Atividade", ""),
        row.get("PV_Texto", ""),
        row.get("Fonte_Arquivo", ""),
    )


def import_rows(rows: list[dict]) -> tuple[int, int, int]:
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Planilha PDS não encontrada: {WORKBOOK}")

    wb = load_workbook(WORKBOOK)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = load_headers(ws)

    source_dates = {
        (
            row["Fonte_Arquivo"],
            row["Data"].strftime("%Y-%m-%d") if hasattr(row["Data"], "strftime") else str(row["Data"]),
        )
        for row in rows
    }
    removed = 0
    for r in range(ws.max_row, 4, -1):
        source = ws.cell(r, headers["Fonte_Arquivo"]).value
        dt = ws.cell(r, headers["Data"]).value
        dt_key = dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt)
        if (source, dt_key) in source_dates:
            ws.delete_rows(r, 1)
            removed += 1

    existing = set()
    date_col = headers.get("Data")
    for r in range(5, ws.max_row + 1):
        current = {}
        for key, col in headers.items():
            current[key] = ws.cell(r, col).value
        if current.get("Data"):
            existing.add(row_key(current))

    inserted = 0
    skipped = 0
    for row in rows:
        if row_key(row) in existing:
            skipped += 1
            continue
        target = ws.max_row + 1
        for key, value in row.items():
            if key in headers:
                ws.cell(target, headers[key]).value = value

        if "MES_REF" in headers:
            ws.cell(target, headers["MES_REF"]).value = f'=IF($A{target}="","",TEXT($A{target},"mmmm/yyyy"))'
        if "MUNICIPIO" in headers:
            ws.cell(target, headers["MUNICIPIO"]).value = f'=IFERROR(XLOOKUP($N{target},Base_Enderecos!$A:$A,Base_Enderecos!$B:$B,""),"")'
        if "SUBPREFEITURA" in headers:
            ws.cell(target, headers["SUBPREFEITURA"]).value = f'=IFERROR(XLOOKUP($N{target},Base_Enderecos!$A:$A,Base_Enderecos!$C:$C,""),"")'
        if "METODO_PADRAO" in headers:
            ws.cell(target, headers["METODO_PADRAO"]).value = '=IF(ISNUMBER(SEARCH("HDD",$F{0})),"HDD",IF(ISNUMBER(SEARCH("VCA",$F{0})),"VCA","MND"))'.format(target)
        if "CHAVE_RELATORIO" in headers:
            ws.cell(target, headers["CHAVE_RELATORIO"]).value = f'=TEXT($A{target},"yyyy-mm")&"|"&$C{target}&"|"&$I{target}'

        existing.add(row_key(row))
        inserted += 1

    wb.save(WORKBOOK)
    root_copy = ROOT / WORKBOOK.name
    shutil.copy2(WORKBOOK, root_copy)
    return inserted, removed, len(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--docx-dir", default=str(DEFAULT_DOCX_DIR), help="Pasta com arquivos Word da PDS")
    parser.add_argument("--all-dates", action="store_true", help="Importa todas as datas do Word, não só a mais recente")
    parser.add_argument("--all-files", action="store_true", help="Lê todos os Word da pasta; por padrão usa só o mais recente")
    args = parser.parse_args()

    docx_dir = Path(args.docx_dir)
    if not docx_dir.exists():
        print(f"AVISO: pasta de PDS Word não encontrada: {docx_dir}")
        return

    files = [p for p in sorted(docx_dir.glob("*.docx")) if not p.name.startswith("~$")]
    if not args.all_files and files:
        files = [max(files, key=lambda p: p.stat().st_mtime)]

    all_rows: list[dict] = []
    for path in files:
        rows = parse_docx(path)
        if not rows:
            continue
        if not args.all_dates:
            latest = max(r["Data"] for r in rows)
            rows = [r for r in rows if r["Data"].date() == latest.date()]
        all_rows.extend(rows)

    inserted, removed, total = import_rows(all_rows)
    by_date = Counter(r["Data"].strftime("%Y-%m-%d") for r in all_rows)
    print("PDS Word importado.")
    print(f"Linhas lidas: {total} | removidas para reimportação: {removed} | inseridas: {inserted}")
    print("Datas:", ", ".join(f"{k}={v}" for k, v in sorted(by_date.items())) or "-")


if __name__ == "__main__":
    main()
