# -*- coding: utf-8 -*-
"""
Gera os JSONs do dashboard antigo a partir das planilhas oficiais.

Fontes principais:
- docs/BASE_DASH_EXTENSAO_POWERBI.xlsx  -> registros do dashboard principal
- docs/EAP_PRODUCAO.xlsx                 -> EAP x Produção e Economias (colunas F/G)

Saídas:
- docs/dados.json
- dados.json
- docs/eap_producao.json
"""
import json
import re
import shutil
import unicodedata
from datetime import datetime, date
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent
DOCS_DIR = ROOT_DIR / "docs"

def escolher_planilha(nome):
    candidatos = [ROOT_DIR / nome, DOCS_DIR / nome]
    existentes = [p for p in candidatos if p.exists()]
    if not existentes:
        raise FileNotFoundError(f"Planilha nao encontrada na raiz nem em docs: {nome}")
    origem = max(existentes, key=lambda p: p.stat().st_mtime)
    for destino in candidatos:
        if destino == origem:
            continue
        if not destino.exists() or origem.stat().st_mtime > destino.stat().st_mtime + 1:
            destino.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(origem, destino)
            print(f"Sincronizado {origem.name}: {origem.parent.name} -> {destino.parent.name}")
    return origem


ARQ_BASE = escolher_planilha("BASE_DASH_EXTENSAO_POWERBI.xlsx")
ARQ_EAP = escolher_planilha("EAP_PRODUCAO.xlsx")

SAIDA_DADOS_DOCS = DOCS_DIR / "dados.json"
SAIDA_DADOS_ROOT = ROOT_DIR / "dados.json"
SAIDA_EAP_DOCS = DOCS_DIR / "eap_producao.json"
SAIDA_EAP_ROOT = ROOT_DIR / "eap_producao.json"

MESES_PT = {
    "janeiro": 1, "jan": 1,
    "fevereiro": 2, "fev": 2,
    "marco": 3, "março": 3, "mar": 3,
    "abril": 4, "abr": 4,
    "maio": 5, "mai": 5,
    "junho": 6, "jun": 6,
    "julho": 7, "jul": 7,
    "agosto": 8, "ago": 8,
    "setembro": 9, "set": 9,
    "outubro": 10, "out": 10,
    "novembro": 11, "nov": 11,
    "dezembro": 12, "dez": 12,
}
MESES_ABREV = {1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez"}
MESES_ORDEM = [MESES_ABREV[i] for i in range(1, 13)]

REQUIRED = {
    "Obra": ["obra"],
    "Bloco": ["bloco"],
    "Tipo": ["tipo_extensao", "tipo extensao", "tipo", "tipoextensao"],
    "Planejado_m": ["extensao_planejada_m", "extensao planejada (m)", "extensao planejada", "planejado_m"],
    "Executado_m": ["extensao_executada_m", "extensao executada (m)", "extensao executada", "executado_m"],
}
OPTIONAL = {
    "Data": ["data"],
    "Status": ["status"],
    "PV": ["pv", "pvs", "qtd pv", "quantidade pv", "qtdpv"],
    "Profundidade_m": ["profundidade_pv_m", "profundidade pv (m)", "profundidade", "profundidade_m"],
    "Economias_Previstas": ["economias prevista", "economias previstas", "economias eap", "econ prev"],
    "Economias_Recebidas": ["economias recebidas", "economias recebida", "econ receb"],
}


def strip_accents(s):
    return unicodedata.normalize("NFD", str(s or "")).encode("ascii", "ignore").decode("ascii")


def norm(s):
    s = strip_accents(s).lower().strip()
    return re.sub(r"[^a-z0-9]", "", s)


def to_float(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    txt = str(v).strip().replace("R$", "").replace(" ", "")
    if "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    try:
        return float(txt)
    except Exception:
        return 0.0


def to_iso_date(v):
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return ""


def normalizar_mes(valor):
    txt = strip_accents(valor).strip().lower()
    if not txt:
        return ""
    if txt.isdigit():
        n = int(txt)
        return MESES_ABREV.get(n, "")
    n = MESES_PT.get(txt) or MESES_PT.get(txt[:3])
    return MESES_ABREV.get(n, str(valor).strip()[:3].title())


def find_col(headers_norm, candidates):
    cand_norm = [norm(c) for c in candidates]
    for i, h in enumerate(headers_norm):
        if h in cand_norm:
            return i
    return None


def detectar_colunas_producao(headers):
    cols = {}
    for idx, h in enumerate(headers):
        hs = strip_accents(h).lower().strip()
        if not hs.startswith("producao"):
            continue
        resto = re.sub(r"^producao\s*", "", hs).strip()
        m = re.search(r"(\d{1,2})\s*[/-]?\s*(\d{4})", resto)
        if m:
            mes, ano = int(m.group(1)), int(m.group(2))
            if 1 <= mes <= 12:
                cols[f"{ano:04d}-{mes:02d}"] = idx
            continue
        m = re.search(r"([a-z]+)\s+(\d{4})", resto)
        if m:
            mes = MESES_PT.get(m.group(1)[:3]) or MESES_PT.get(m.group(1))
            ano = int(m.group(2))
            if mes:
                cols[f"{ano:04d}-{mes:02d}"] = idx
    return dict(sorted(cols.items()))


def ler_registros_base():
    if not ARQ_BASE.exists():
        raise FileNotFoundError(f"Planilha base não encontrada: {ARQ_BASE}")

    wb = load_workbook(ARQ_BASE, data_only=True, read_only=True)
    ws = wb["BASE_DASH_EXTENSAO"] if "BASE_DASH_EXTENSAO" in wb.sheetnames else wb[wb.sheetnames[0]]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers_norm = [norm(h) for h in headers]
    prod_cols = detectar_colunas_producao(headers)

    col_index = {}
    for out_name, candidates in REQUIRED.items():
        found = find_col(headers_norm, candidates)
        if found is None:
            raise ValueError(f"Coluna obrigatória ausente na base: {out_name}")
        col_index[out_name] = found
    for out_name, candidates in OPTIONAL.items():
        found = find_col(headers_norm, candidates)
        if found is not None:
            col_index[out_name] = found

    registros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue

        def get_idx(name):
            i = col_index.get(name)
            return row[i] if i is not None and i < len(row) else None

        registros.append({
            "Data": to_iso_date(get_idx("Data")),
            "Status": str(get_idx("Status") or "").strip(),
            "Obra": str(get_idx("Obra") or "").strip(),
            "Bloco": str(get_idx("Bloco") or "").strip(),
            "Tipo": str(get_idx("Tipo") or "").strip(),
            "Planejado_m": to_float(get_idx("Planejado_m")),
            "Executado_m": to_float(get_idx("Executado_m")),
            "PV": to_float(get_idx("PV")),
            "Profundidade_m": to_float(get_idx("Profundidade_m")),
            "Economias_Previstas": to_float(get_idx("Economias_Previstas")),
            "Economias_Recebidas": to_float(get_idx("Economias_Recebidas")),
            "ProducaoMensal": {k: to_float(row[idx]) for k, idx in prod_cols.items() if idx < len(row) and row[idx] not in (None, "")},
        })

    return registros, list(prod_cols.keys())


def localizar_header(ws, obrigatorias):
    obrig_norm = {norm(x) for x in obrigatorias}
    for row_idx in range(1, min(ws.max_row, 15) + 1):
        vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        vals_norm = [norm(v) for v in vals]
        if obrig_norm.issubset(set(vals_norm)):
            return row_idx, vals, vals_norm
    return None, [], []


def ler_eap_producao():
    if not ARQ_EAP.exists():
        raise FileNotFoundError(f"Planilha EAP não encontrada: {ARQ_EAP}")

    wb = load_workbook(ARQ_EAP, data_only=True, read_only=True)
    ws = wb["EAP_PRODUCAO"] if "EAP_PRODUCAO" in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row, headers, headers_norm = localizar_header(ws, ["Ano", "Mes", "EAP", "Produzido"])
    if not header_row:
        raise ValueError("Não encontrei o cabeçalho Ano/Mes/EAP/Produzido em EAP_PRODUCAO.xlsx")

    cols = {
        "ano": find_col(headers_norm, ["Ano"]),
        "mes": find_col(headers_norm, ["Mes", "Mês"]),
        "eap": find_col(headers_norm, ["EAP"]),
        "produzido": find_col(headers_norm, ["Produzido"]),
        "economias_eap": find_col(headers_norm, ["Economias EAP", "Economias Previstas", "Economias prevista"]),
        "economias_recebidas": find_col(headers_norm, ["Economias Recebidas", "Economias recebida"]),
    }

    mensal = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        ano = row[cols["ano"]] if cols["ano"] is not None and cols["ano"] < len(row) else None
        mes = row[cols["mes"]] if cols["mes"] is not None and cols["mes"] < len(row) else None
        if ano in (None, "") and mes in (None, ""):
            continue
        ano_i = int(to_float(ano))
        if not ano_i:
            continue
        mes_txt = normalizar_mes(mes)
        item = {
            "ano": ano_i,
            "mes": mes_txt,
            "eap": round(to_float(row[cols["eap"]]), 6),
            "produzido": round(to_float(row[cols["produzido"]]), 6),
            "economias_eap": round(to_float(row[cols["economias_eap"]]) if cols["economias_eap"] is not None else 0, 6),
            "economias_recebidas": round(to_float(row[cols["economias_recebidas"]]) if cols["economias_recebidas"] is not None else 0, 6),
        }
        item["saldo_mes"] = round(item["produzido"] - item["eap"], 6)
        item["saldo_economias"] = round(item["economias_recebidas"] - item["economias_eap"], 6)
        mensal.append(item)

    ordem = {m: i for i, m in enumerate(MESES_ORDEM)}
    mensal.sort(key=lambda x: (x["ano"], ordem.get(x["mes"], 99)))

    cards = {}
    econ_cards = {}
    saldo_acum_por_ano = {}
    for item in mensal:
        ano = str(item["ano"])
        cards.setdefault(ano, {"eap": 0.0, "produzido": 0.0, "saldo": 0.0})
        econ_cards.setdefault(ano, {"economias_eap": 0.0, "economias_recebidas": 0.0, "saldo": 0.0})
        saldo_acum_por_ano[ano] = saldo_acum_por_ano.get(ano, 0.0) + item["saldo_mes"]
        item["saldo_acum"] = round(saldo_acum_por_ano[ano], 6)
        cards[ano]["eap"] += item["eap"]
        cards[ano]["produzido"] += item["produzido"]
        econ_cards[ano]["economias_eap"] += item["economias_eap"]
        econ_cards[ano]["economias_recebidas"] += item["economias_recebidas"]

    for ano, c in cards.items():
        c["eap"] = round(c["eap"], 6)
        c["produzido"] = round(c["produzido"], 6)
        c["saldo"] = round(c["produzido"] - c["eap"], 6)
    for ano, c in econ_cards.items():
        c["economias_eap"] = round(c["economias_eap"], 6)
        c["economias_recebidas"] = round(c["economias_recebidas"], 6)
        c["saldo"] = round(c["economias_recebidas"] - c["economias_eap"], 6)

    total_econ = {
        "economias_eap": round(sum(c["economias_eap"] for c in econ_cards.values()), 6),
        "economias_recebidas": round(sum(c["economias_recebidas"] for c in econ_cards.values()), 6),
    }
    total_econ["saldo"] = round(total_econ["economias_recebidas"] - total_econ["economias_eap"], 6)

    return {
        "atualizado_em": datetime.now(ZoneInfo("America/Sao_Paulo")).isoformat(),
        "fonte": ARQ_EAP.name,
        "aba": ws.title,
        "meses": MESES_ORDEM,
        "cards": cards,
        "mensal": mensal,
        "eap_economias": {
            "cards": econ_cards,
            "total": total_econ,
            "mensal": [
                {
                    "ano": i["ano"],
                    "mes": i["mes"],
                    "economias_eap": i["economias_eap"],
                    "economias_recebidas": i["economias_recebidas"],
                    "saldo": i["saldo_economias"],
                }
                for i in mensal
            ],
        },
    }


def aplicar_producao_base_na_eap(eap, registros):
    """Mantem o planejado da EAP, mas usa a producao mensal da base oficial."""
    produzido_por_mes = {}
    for registro in registros:
        for ym, valor in (registro.get("ProducaoMensal") or {}).items():
            produzido_por_mes[ym] = produzido_por_mes.get(ym, 0.0) + to_float(valor)

    mes_num_por_abrev = {v: k for k, v in MESES_ABREV.items()}
    cards = {}
    saldo_acum_por_ano = {}

    for item in eap.get("mensal", []):
        mes_num = mes_num_por_abrev.get(item.get("mes"))
        ym = f"{int(item['ano']):04d}-{mes_num:02d}" if mes_num else ""
        item["produzido"] = round(produzido_por_mes.get(ym, 0.0), 6)
        item["saldo_mes"] = round(item["produzido"] - item["eap"], 6)

        ano = str(item["ano"])
        saldo_acum_por_ano[ano] = saldo_acum_por_ano.get(ano, 0.0) + item["saldo_mes"]
        item["saldo_acum"] = round(saldo_acum_por_ano[ano], 6)

        cards.setdefault(ano, {"eap": 0.0, "produzido": 0.0, "saldo": 0.0})
        cards[ano]["eap"] += item["eap"]
        cards[ano]["produzido"] += item["produzido"]

    for card in cards.values():
        card["eap"] = round(card["eap"], 6)
        card["produzido"] = round(card["produzido"], 6)
        card["saldo"] = round(card["produzido"] - card["eap"], 6)

    eap["cards"] = cards
    eap["fonte_produzido"] = ARQ_BASE.name
    eap["regra_produzido"] = "Colunas ProducaoMensal da BASE_DASH_EXTENSAO_POWERBI.xlsx"
    return eap


def salvar_json(caminho, conteudo):
    caminho.parent.mkdir(parents=True, exist_ok=True)
    with caminho.open("w", encoding="utf-8") as f:
        json.dump(conteudo, f, ensure_ascii=False, indent=2)


def main():
    registros, meses_producao = ler_registros_base()
    eap = ler_eap_producao()
    eap = aplicar_producao_base_na_eap(eap, registros)
    payload = {
        "atualizado_em": datetime.now(ZoneInfo("America/Sao_Paulo")).isoformat(),
        "fonte_registros": ARQ_BASE.name,
        "fonte_eap_economias": ARQ_EAP.name,
        "meses_producao": meses_producao,
        "registros": registros,
        "eap_producao": eap,
    }
    salvar_json(SAIDA_DADOS_DOCS, payload)
    salvar_json(SAIDA_DADOS_ROOT, payload)
    salvar_json(SAIDA_EAP_DOCS, eap)
    salvar_json(SAIDA_EAP_ROOT, eap)

    total = eap["eap_economias"]["total"]
    print("OK: dados.json e eap_producao.json atualizados.")
    print(f"Registros: {len(registros)}")
    print(f"Economias EAP/Previstas: {total['economias_eap']:.0f}")
    print(f"Economias Recebidas/Realizadas: {total['economias_recebidas']:.0f}")
    print(f"Saldo Economias: {total['saldo']:.0f}")


if __name__ == "__main__":
    main()
