# -*- coding: utf-8 -*-
import json
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DOCS = ROOT / "docs"
XLSX_NAME = "pds_word_completo_preenchido(2).xlsx"
TARGET_DATES = {"2026-06-12", "2026-06-13", "2026-06-15"}


def pv_da_atividade(atividade):
    m = re.search(r"\bPV(?!s\b)[-\s]*([A-Z]?\d+(?:\.\d+)?|[A-Z])", atividade, re.I)
    return m.group(1) if m else ""


def registro(data, obra, equipe, atividade):
    pv = pv_da_atividade(atividade)
    trecho = ""
    m = re.search(r"\bPV[-\s]*([A-Z]?\d+(?:\.\d+)?|[A-Z])\s*(?:ao|a|até|-)\s*PV[-\s]*([A-Z]?\d+(?:\.\d+)?|[A-Z])", atividade, re.I)
    if m:
        trecho = f"{m.group(1)}-{m.group(2)}"
        pv = m.group(1)
    return {
        "Data": data,
        "Obra": obra,
        "Equipe": equipe,
        "Atividade": atividade,
        "PV_Inicio": pv,
        "PV_Fim": "",
        "PV_Texto": pv,
        "Trecho": trecho,
    }


NOVOS = [
    registro("2026-06-12", "CT JOÃO CANZI", "Márcio", "Transformação PV 19"),
    registro("2026-06-12", "CT JOÃO CANZI", "Bruno", "Transformação PV 11"),
    registro("2026-06-12", "RCE ELVIRA", "Valter", "VCA, PVs e Ligações"),
    registro("2026-06-12", "INTERL. AYRTON SENNA", "Cidão", "VCA, PVs e Ligações"),
    registro("2026-06-12", "RCE CONJ. UNIÃO", "Jhonatan", "Transformação PV 48"),
    registro("2026-06-12", "RCE ELIAS SLEIMAN", "Edvando", "VCA PV 23 ao PV 24"),
    registro("2026-06-12", "RCE AMÉRICO MACHADO/ELIS REGINA", "Cesar", "VCA PV 01 ao PV 02"),
    registro("2026-06-12", "CTS LOURDES Complementar", "Marcelo", "Interligação entre PV-05 e PV-E"),
    registro("2026-06-12", "RCE SÃO LUCAS", "Ricardo (PI 10 ao PI 11)", "VCA, PVs e Ligações"),
    registro("2026-06-12", "RCE SÃO LUCAS", "Claudinei (PI 38 ao PI 81)", "VCA, PVs e Ligações"),
    registro("2026-06-12", "RCE SÃO LUCAS", "Miro (PI 61 ao PI 62)", "VCA, PVs e Ligações"),
    registro("2026-06-12", "RCE SÃO LUCAS", "Emanuel (PI 25 ao PI 26)", "VCA, PVs e Ligações"),
    registro("2026-06-12", "RCE SÃO LUCAS", "Antonio (PI 55 ao PI 56)", "VCA, PVs e Ligações"),
    registro("2026-06-12", "SERVIÇOS REPOSIÇÃO", "Leandro", "Reparos Gerais"),
    registro("2026-06-12", "Vacal: João Canzi", "Henko", "Apoio às equipes"),
    registro("2026-06-12", "Vacal: João Canzi", "Henko", "Apoio às equipes"),
    registro("2026-06-12", "Vacal: ITI 15", "Renasce", "Apoio às equipes"),
    registro("2026-06-12", "Guindauto", "Marcio", "Apoio as equipes"),
    registro("2026-06-12", "Guindauto", "Luiz", "Apoio as equipes"),
    registro("2026-06-13", "CT JOÃO CANZI", "Cesar", "Transformação PV 11"),
    registro("2026-06-13", "RCE ELVIRA", "Valter", "VCA, PVs e Ligações"),
    registro("2026-06-13", "INTERL. AYRTON SENNA", "Cidão", "VCA, PVs e Ligações"),
    registro("2026-06-13", "RCE CONJ. UNIÃO", "Jhonatan", "Transformação PV 48"),
    registro("2026-06-13", "RCE ELIAS SLEIMAN", "Edvando", "VCA PV 23 ao PV 24"),
    registro("2026-06-13", "CTS LOURDES Complementar", "Marcelo", "Interligação entre PV-05 e PV-E"),
    registro("2026-06-13", "RCE SÃO LUCAS", "Ricardo (PI 10 ao PI 11)", "VCA, PVs e Ligações"),
    registro("2026-06-13", "RCE SÃO LUCAS", "Claudinei (PI 38 ao PI 81)", "VCA, PVs e Ligações"),
    registro("2026-06-13", "RCE SÃO LUCAS", "Miro (PI 61 ao PI 62)", "VCA, PVs e Ligações"),
    registro("2026-06-13", "RCE SÃO LUCAS", "Emanuel (PI 25 ao PI 26)", "VCA, PVs e Ligações"),
    registro("2026-06-13", "RCE SÃO LUCAS", "Antonio (PI 55 ao PI 56)", "VCA, PVs e Ligações"),
    registro("2026-06-13", "SERVIÇOS REPOSIÇÃO", "Leandro", "Reparos Gerais"),
    registro("2026-06-13", "Vacal: ITI-15", "Henko", "Apoio às equipes"),
    registro("2026-06-13", "Vacal: João Canzi", "Henko", "Apoio às equipes"),
    registro("2026-06-13", "Vacal: João Canzi", "Henko", "Apoio às equipes"),
    registro("2026-06-13", "Guindauto", "Marcio", "Apoio as equipes"),
    registro("2026-06-13", "Guindauto", "Luiz", "Apoio as equipes"),
    registro("2026-06-15", "CT JOÃO CANZI", "Cesar", "Transformação PV 11"),
    registro("2026-06-15", "RCE ELVIRA", "Valter", "VCA, PVs e Ligações"),
    registro("2026-06-15", "INTERL. AYRTON SENNA", "Cidão", "VCA, PVs e Ligações"),
    registro("2026-06-15", "RCE CONJ. UNIÃO", "Jhonatan", "Transformação PV 48"),
    registro("2026-06-15", "RCE ELIAS SLEIMAN", "Edvando", "VCA PV 24 ao PV 25"),
    registro("2026-06-15", "CTS LOURDES Complementar", "Marcelo", "Interligação entre PV-05 e PV-E"),
    registro("2026-06-15", "RCE RAULZITO", "Bruno", "VCA, PVs e Ligações"),
    registro("2026-06-15", "RCE RAULZITO", "Medeiros", "VCA, PVs e Ligações"),
    registro("2026-06-15", "RCE SÃO LUCAS", "Ricardo (PI 10 ao PI 11)", "VCA, PVs e Ligações"),
    registro("2026-06-15", "RCE SÃO LUCAS", "Claudinei (PI 38 ao PI 81)", "VCA, PVs e Ligações"),
    registro("2026-06-15", "RCE SÃO LUCAS", "Miro (PI 61 ao PI 62)", "VCA, PVs e Ligações"),
    registro("2026-06-15", "RCE SÃO LUCAS", "Emanuel (PI 25 ao PI 26)", "VCA, PVs e Ligações"),
    registro("2026-06-15", "RCE SÃO LUCAS", "Antonio (PI 55 ao PI 56)", "VCA, PVs e Ligações"),
    registro("2026-06-15", "SERVIÇOS REPOSIÇÃO", "Leandro", "Reparos Gerais"),
    registro("2026-06-15", "Vacal: João Canzi", "Henko", "Apoio às equipes"),
    registro("2026-06-15", "Vacal: João Canzi", "Henko", "Apoio às equipes"),
    registro("2026-06-15", "Vacal: ITI 15", "Henko", "Apoio às equipes"),
    registro("2026-06-15", "Guindauto", "Marcio", "Apoio as equipes"),
    registro("2026-06-15", "Guindauto", "Luiz", "Apoio as equipes"),
]


def normalizar_data(valor):
    if hasattr(valor, "date"):
        return valor.date().isoformat()
    return str(valor or "").strip()[:10]


def atualizar_xlsx(caminho):
    wb = load_workbook(caminho)
    ws = wb["PDS_Lancamentos"] if "PDS_Lancamentos" in wb.sheetnames else wb[wb.sheetnames[0]]

    header_row = None
    headers = []
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        vals = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if {"Data", "Obra", "Equipe", "Atividade"}.issubset(set(vals)):
            header_row = row_idx
            headers = [str(v or "").strip() for v in vals]
            break
    if not header_row:
        raise RuntimeError("Cabeçalho do PDS não encontrado")

    col_data = headers.index("Data") + 1
    for row_idx in range(ws.max_row, header_row, -1):
        if normalizar_data(ws.cell(row_idx, col_data).value) in TARGET_DATES:
            ws.delete_rows(row_idx)

    colunas = {name: headers.index(name) + 1 for name in headers if name}
    for item in NOVOS:
        row_idx = ws.max_row + 1
        for campo, valor in item.items():
            if campo in colunas:
                ws.cell(row_idx, colunas[campo]).value = valor
    wb.save(caminho)


def atualizar_json(caminho):
    atuais = json.loads(caminho.read_text(encoding="utf-8")) if caminho.exists() else []
    atuais = [r for r in atuais if str(r.get("data")) not in TARGET_DATES]
    novos = [
        {
            "data": r["Data"],
            "obra": r["Obra"],
            "equipe": r["Equipe"],
            "atividade": r["Atividade"],
            "trecho": r["Trecho"],
            "pv": r["PV_Texto"],
        }
        for r in NOVOS
    ]
    atuais.extend(novos)
    atuais.sort(key=lambda r: (str(r.get("data", "")), str(r.get("obra", "")), str(r.get("equipe", ""))))
    caminho.write_text(json.dumps(atuais, ensure_ascii=False, indent=2), encoding="utf-8")


def main():
    docs_xlsx = DOCS / XLSX_NAME
    root_xlsx = ROOT / XLSX_NAME
    if not docs_xlsx.exists() and root_xlsx.exists():
        shutil.copy2(root_xlsx, docs_xlsx)
    if not docs_xlsx.exists():
        raise FileNotFoundError(docs_xlsx)

    atualizar_xlsx(docs_xlsx)
    shutil.copy2(docs_xlsx, root_xlsx)
    atualizar_json(DOCS / "pds_data.json")
    shutil.copy2(DOCS / "pds_data.json", ROOT / "pds_data.json")
    print(f"PDS atualizados: {len(NOVOS)} registros em {', '.join(sorted(TARGET_DATES))}")


if __name__ == "__main__":
    main()
