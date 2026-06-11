# -*- coding: utf-8 -*-
"""Inclui o plano de furo CTS João Canzi PV-10 ao PV-11 na base do mapa."""

from __future__ import annotations

import math
import shutil
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
FILES = [ROOT / "planilha_base_mapa.xlsx", ROOT / "docs" / "planilha_base_mapa.xlsx"]


def header_map(ws):
    return {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}


def get_point(ws, headers, point_id):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, headers["obra_id"]).value != "cts_joao_canzi":
            continue
        if ws.cell(row, headers["ponto_id"]).value == point_id:
            return {name: ws.cell(row, col).value for name, col in headers.items()}
    raise ValueError(f"Ponto não encontrado: {point_id}")


def upsert_trecho(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    pontos = wb["PONTOS"]
    trechos = wb["TRECHOS"]
    hp = header_map(pontos)
    ht = header_map(trechos)

    pv10 = get_point(pontos, hp, "JC-PV-10")
    pv11 = get_point(pontos, hp, "JC-PV-11")

    target_row = None
    for row in range(2, trechos.max_row + 1):
        if trechos.cell(row, ht["trecho_id"]).value == "JC-005":
            target_row = row
            break
        if (
            trechos.cell(row, ht["obra_id"]).value == "cts_joao_canzi"
            and trechos.cell(row, ht["pv_inicio"]).value == "PV-10"
            and trechos.cell(row, ht["pv_fim"]).value == "PV-11"
        ):
            target_row = row
            break

    if target_row is None:
        target_row = trechos.max_row + 1

    extensao = round(
        math.dist(
            (float(pv10["coord_n"]), float(pv10["coord_e"])),
            (float(pv11["coord_n"]), float(pv11["coord_e"])),
        ),
        3,
    )
    prof_inicial = float(pv10["profundidade_m"])
    prof_final = float(pv11["profundidade_m"])

    values = {
        "obra_id": "cts_joao_canzi",
        "trecho_id": "JC-005",
        "trecho_nome": "PV-10 - PV-11",
        "pv_inicio": "PV-10",
        "pv_fim": "PV-11",
        "material": "PEAD LISO",
        "dn": 355,
        "metodo": "Metodo Nao Destrutivo - MND",
        "extensao_m": extensao,
        "prof_inicial_m": prof_inicial,
        "prof_final_m": prof_final,
        "prof_max_m": max(prof_inicial, prof_final),
        "status": "em andamento",
        "observacao": "Plano de furo FL-05; PV-10 ao PV-11; arquivo: planos_furo/cts_joao_canzi_fl05_pv10_pv11.pdf.",
        "lat_inicio": pv10["latitude"],
        "lon_inicio": pv10["longitude"],
        "lat_fim": pv11["latitude"],
        "lon_fim": pv11["longitude"],
    }

    for name, value in values.items():
        trechos.cell(target_row, ht[name]).value = value

    wb.save(path)
    wb.close()
    print(f"{path.name}: trecho JC-005 salvo na linha {target_row} com {extensao} m")


def main() -> None:
    for path in FILES:
        upsert_trecho(path)

    root_pdf = ROOT / "planos_furo" / "cts_joao_canzi_fl05_pv10_pv11.pdf"
    docs_pdf = ROOT / "docs" / "planos_furo" / "cts_joao_canzi_fl05_pv10_pv11.pdf"
    if root_pdf.exists():
        docs_pdf.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(root_pdf, docs_pdf)


if __name__ == "__main__":
    main()
