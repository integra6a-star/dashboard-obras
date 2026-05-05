# -*- coding: utf-8 -*-
"""
Gera docs/medicao.json a partir de uma planilha Excel (docs/medicao.xlsx).

Layout esperado (aba "medição" ou "medicao"):
Coluna A: DATA (ex: "jun/25", "jul/25", ou data)
Coluna B: MEDIÇÃO (valor)
Coluna C: AMORTIZAÇÃO (valor)

Obs.: TOTAL CONTRATO e ADIANTAMENTO podem ficar como constantes abaixo.
"""
import os
import sys
import json
import re
from datetime import datetime
from openpyxl import load_workbook

# ========= CONFIG =========
PASTA_RAIZ = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
PASTA_DOCS = os.path.join(PASTA_RAIZ, "docs")

# nomes candidatos do arquivo
CANDIDATOS_XLSX = [
    "medicao.xlsx",
    "medição.xlsx",
    "MEDICAO.xlsx",
    "MEDIÇÃO.xlsx",
]

ARQ_SAIDA = os.path.join(PASTA_DOCS, "medicao.json")

# >>> AJUSTE AQUI (se quiser):
TOTAL_CONTRATO_PADRAO = 263_030_690.00
ADIANTAMENTO_PADRAO   = 52_606_138.00
# ==========================

MESES_PT = {
    1:"janeiro",2:"fevereiro",3:"março",4:"abril",5:"maio",6:"junho",
    7:"julho",8:"agosto",9:"setembro",10:"outubro",11:"novembro",12:"dezembro"
}

def now_iso_local():
    return datetime.now().astimezone().isoformat()

def encontrar_excel():
    # 1) tenta candidatos fixos
    for nome in CANDIDATOS_XLSX:
        p = os.path.join(PASTA_DOCS, nome)
        if os.path.exists(p):
            return p

    # 2) fallback: qualquer xlsx que tenha "medic" no nome
    try:
        for nome in os.listdir(PASTA_DOCS):
            if nome.lower().endswith(".xlsx") and "medic" in nome.lower():
                return os.path.join(PASTA_DOCS, nome)
    except Exception:
        pass
    return None

def normalizar_header(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = (s.replace("ç","c")
           .replace("á","a").replace("à","a").replace("ã","a").replace("â","a")
           .replace("é","e").replace("ê","e")
           .replace("í","i")
           .replace("ó","o").replace("ô","o").replace("õ","o")
           .replace("ú","u"))
    s = re.sub(r"\s+", " ", s)
    return s

def to_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return 0.0
    # remove moeda e espaços
    s = s.replace("R$", "").replace(" ", "")
    # troca milhar/decimal pt-BR
    # se tiver ambos '.' e ',', assume '.' milhar e ',' decimal
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        # se só vírgula, é decimal
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0

def parse_mes(value):
    """
    Aceita:
    - datetime/date
    - "jun/25", "jun-25", "jun25"
    - "2025-06-01"
    Retorna: (ano, mes) ou None
    """
    if value is None:
        return None

    if hasattr(value, "year") and hasattr(value, "month"):
        return int(value.year), int(value.month)

    s = str(value).strip()
    if s == "":
        return None

    # ISO-like
    m = re.match(r"^(\d{4})-(\d{2})", s)
    if m:
        return int(m.group(1)), int(m.group(2))

    # "jun/25"
    s2 = s.lower().replace(" ", "")
    s2 = s2.replace("-", "/")
    m = re.match(r"^([a-zç]{3})\/?(\d{2,4})$", s2)
    if m:
        mon = m.group(1)
        yy = m.group(2)
        mon_map = {
            "jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,
            "jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12
        }
        if mon not in mon_map:
            return None
        year = int(yy)
        if year < 100:
            year = 2000 + year
        return year, mon_map[mon]

    # "06/2025" ou "6/2025"
    m = re.match(r"^(\d{1,2})\/(\d{4})$", s2)
    if m:
        return int(m.group(2)), int(m.group(1))

    return None

def label_mes(ano, mes):
    return f"{MESES_PT.get(mes, str(mes))} de {ano}"

def ler_planilha(caminho):
    wb = load_workbook(caminho, data_only=True)

    # tenta achar aba
    sheet = None
    for name in wb.sheetnames:
        nn = normalizar_header(name)
        if nn in ("medicao", "medicao ", "medicao", "medição", "medicao-"):
            sheet = wb[name]; break
    if sheet is None:
        # tenta por contém
        for name in wb.sheetnames:
            if "medic" in normalizar_header(name):
                sheet = wb[name]; break
    if sheet is None:
        sheet = wb.active

    ws = sheet

    # detectar cabeçalho na linha 1..5
    header_row = None
    cols = {}
    for r in range(1, 6):
        a = normalizar_header(ws.cell(r, 1).value)
        b = normalizar_header(ws.cell(r, 2).value)
        c = normalizar_header(ws.cell(r, 3).value)
        if ("data" in a) and ("med" in b) and ("amor" in c):
            header_row = r
            cols = {"data":1, "medicao":2, "amortizacao":3}
            break

    # fallback: assume A,B,C com header na linha 1
    if header_row is None:
        header_row = 1
        cols = {"data":1, "medicao":2, "amortizacao":3}

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        raw_data = ws.cell(r, cols["data"]).value
        ym = parse_mes(raw_data)
        med = to_float(ws.cell(r, cols["medicao"]).value)
        amo = to_float(ws.cell(r, cols["amortizacao"]).value)

        # linha vazia de verdade -> para
        if ym is None and med == 0 and amo == 0:
            continue

        if ym is None:
            # ignora linhas sem data
            continue

        ano, mes = ym
        rows.append({
            "ano": ano,
            "mes": mes,
            "data_label": label_mes(ano, mes),
            "ym": f"{ano:04d}-{mes:02d}",
            "medicao": med,
            "amortizacao": amo
        })

    # ordena por ano-mes
    rows.sort(key=lambda x: x["ym"])
    return rows

def build_series(rows, campo, total_base):
    """
    campo: 'medicao' ou 'amortizacao'
    total_base: total contrato / adiantamento
    Retorna: (serie, total, pct, saldo_final)
    """
    acum = 0.0
    serie = []
    total = sum(float(r.get(campo, 0) or 0) for r in rows)
    for r in rows:
        v = float(r.get(campo, 0) or 0)
        acum += v
        saldo = float(total_base) - acum
        serie.append({
            "ym": r["ym"],
            "data": r["data_label"],
            "valor": v,
            "acumulado": acum,
            "saldo": saldo
        })
    pct = (total / total_base * 100.0) if total_base and total_base > 0 else 0.0
    saldo_final = float(total_base) - float(total)
    return serie, total, pct, saldo_final

def main():
    xlsx = encontrar_excel()
    if not xlsx:
        print("Arquivo não encontrado dentro de:", PASTA_DOCS)
        print("Esperado um desses:", ", ".join(CANDIDATOS_XLSX))
        sys.exit(1)

    rows = ler_planilha(xlsx)

    total_contrato = TOTAL_CONTRATO_PADRAO
    adiantamento = ADIANTAMENTO_PADRAO

    serie_med, total_medido, pct_faturado, saldo_contratual = build_series(rows, "medicao", total_contrato)
    serie_amo, total_amort, pct_amort, saldo_amort = build_series(rows, "amortizacao", adiantamento)

    out = {
        "atualizado_em": now_iso_local(),
        "fonte_excel": os.path.basename(xlsx),
        "contrato_total": float(total_contrato),
        "adiantamento_total": float(adiantamento),
        "medicoes": {
            "total_contrato": float(total_contrato),
            "total_medido": float(total_medido),
            "saldo_contratual": float(saldo_contratual),
            "pct_faturado": float(pct_faturado),
            "serie": serie_med
        },
        "amortizacao": {
            "adiantamento": float(adiantamento),
            "total_amortizado": float(total_amort),
            "saldo_amortizacao": float(saldo_amort),
            "pct_amortizado": float(pct_amort),
            "serie": serie_amo
        }
    }

    os.makedirs(PASTA_DOCS, exist_ok=True)
    with open(ARQ_SAIDA, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"OK: {ARQ_SAIDA} gerado com {len(rows)} linhas.")
    print("Excel lido de:", xlsx)
    print(f"Resumo: total_medido={total_medido:.2f} | pct_faturado={pct_faturado:.2f}% | total_amortizado={total_amort:.2f} | pct_amortizado={pct_amort:.2f}%")

if __name__ == "__main__":
    main()
