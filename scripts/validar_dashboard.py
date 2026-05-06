# -*- coding: utf-8 -*-
"""Gera conferência automática entre planilhas e JSONs do dashboard."""

from __future__ import annotations

import json
import math
import re
import shutil
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
BASE_XLSX = ROOT / "BASE_DASH_EXTENSAO_POWERBI.xlsx"
MAPA_XLSX = ROOT / "planilha_base_mapa.xlsx"
PDS_JSON = DOCS / "pds_data.json"
DE_PARA = ROOT / "de_para_obras.json"
OUT_JSON = ROOT / "validacao_dashboard.json"
OUT_HIST = ROOT / "historico_atualizacoes.json"
OUT_TXT = ROOT / "relatorio_atualizacao.txt"


def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def norm(text: str) -> str:
    text = unicodedata.normalize("NFKD", str(text or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def num(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def load_de_para() -> dict[str, str]:
    if not DE_PARA.exists():
        return {}
    raw = json.loads(DE_PARA.read_text(encoding="utf-8"))
    return {norm(k): v for k, v in raw.items()}


def canonical(name: str, aliases: dict[str, str]) -> str:
    key = norm(name)
    return aliases.get(key, str(name or "").strip() or "Sem obra")


def read_base(aliases: dict[str, str]) -> dict:
    wb = load_workbook(BASE_XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    col_obra = headers.get("Obra", 1)
    col_plan = headers.get("Extensao_Planej", 5)
    col_exec = headers.get("Extensao_Execut", 6)

    by_obra = defaultdict(lambda: {"planejado": 0.0, "executado": 0.0, "linhas": 0, "origens": set()})
    for r in range(2, ws.max_row + 1):
        obra_raw = ws.cell(r, col_obra).value
        if not obra_raw:
            continue
        obra = canonical(str(obra_raw), aliases)
        planejado = num(ws.cell(r, col_plan).value)
        executado = num(ws.cell(r, col_exec).value)
        by_obra[obra]["planejado"] += planejado
        by_obra[obra]["executado"] += executado
        by_obra[obra]["linhas"] += 1
        by_obra[obra]["origens"].add(str(obra_raw).strip())

    return {
        "planejado": sum(v["planejado"] for v in by_obra.values()),
        "executado": sum(v["executado"] for v in by_obra.values()),
        "obras": by_obra,
    }


def find_header_row(ws, required: str) -> tuple[int, dict[str, int]]:
    for row in range(1, min(ws.max_row, 20) + 1):
        headers = {str(ws.cell(row, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(row, c).value}
        if required in headers:
            return row, headers
    return 1, {}


def read_mapa(aliases: dict[str, str]) -> dict:
    wb = load_workbook(MAPA_XLSX, read_only=True, data_only=True)
    ws = wb["TRECHOS"] if "TRECHOS" in wb.sheetnames else wb.active
    header_row, headers = find_header_row(ws, "obra_id")
    col_obra = headers.get("obra_id", 1)
    col_ext = headers.get("extensao_m", 9)
    col_status = headers.get("status")
    col_mat = headers.get("material")
    col_dn = headers.get("dn")

    by_obra = defaultdict(lambda: {"executado_mapa": 0.0, "linhas": 0, "materiais": defaultdict(float), "origens": set()})
    total = 0.0
    for r in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(r, col_obra).value
        ext = num(ws.cell(r, col_ext).value)
        if not raw and not ext:
            continue
        status = str(ws.cell(r, col_status).value or "").strip() if col_status else ""
        obra = canonical(str(raw or "Sem obra"), aliases)
        mat = str(ws.cell(r, col_mat).value or "Sem material").strip() if col_mat else "Sem material"
        dn = str(ws.cell(r, col_dn).value or "Sem diâmetro").strip() if col_dn else "Sem diâmetro"
        by_obra[obra]["executado_mapa"] += ext
        by_obra[obra]["linhas"] += 1
        by_obra[obra]["materiais"][f"{mat} | {dn}"] += ext
        by_obra[obra]["origens"].add(str(raw or "").strip())
        total += ext

    return {"executado_mapa": total, "obras": by_obra}


def read_pds() -> dict:
    if not PDS_JSON.exists():
        return {"total": 0, "ultima_data": "", "hoje": 0, "por_obra_ultima_data": []}
    data = json.loads(PDS_JSON.read_text(encoding="utf-8"))
    today = datetime.now().strftime("%Y-%m-%d")
    dates = sorted({str(r.get("data", "")) for r in data if r.get("data")})
    last_date = dates[-1] if dates else ""
    today_count = sum(1 for r in data if r.get("data") == today)
    per_last = defaultdict(int)
    for row in data:
        if row.get("data") == last_date:
            per_last[row.get("obra") or "Sem obra"] += 1
    return {
        "total": len(data),
        "ultima_data": last_date,
        "hoje": today_count,
        "por_obra_ultima_data": [{"obra": k, "atividades": v} for k, v in sorted(per_last.items(), key=lambda x: (-x[1], x[0]))],
    }


def file_info(path: Path) -> dict:
    if not path.exists():
        return {"arquivo": path.name, "existe": False}
    stat = path.stat()
    return {
        "arquivo": path.name,
        "existe": True,
        "tamanho_kb": round(stat.st_size / 1024, 1),
        "modificado": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
    }


def build_validation() -> dict:
    aliases = load_de_para()
    base = read_base(aliases)
    mapa = read_mapa(aliases)
    pds = read_pds()

    obras = sorted(set(base["obras"].keys()) | set(mapa["obras"].keys()))
    divergencias = []
    saldo_por_obra = []
    for obra in obras:
        b = base["obras"].get(obra, {})
        m = mapa["obras"].get(obra, {})
        planejado = float(b.get("planejado", 0.0))
        executado_base = float(b.get("executado", 0.0))
        executado_mapa = float(m.get("executado_mapa", 0.0))
        diff = executado_base - executado_mapa
        saldo = planejado - executado_base
        item = {
            "obra": obra,
            "planejado": round(planejado, 2),
            "executado_base": round(executado_base, 2),
            "executado_mapa": round(executado_mapa, 2),
            "diferenca_base_mapa": round(diff, 2),
            "saldo": round(saldo, 2),
            "linhas_base": int(b.get("linhas", 0)),
            "linhas_mapa": int(m.get("linhas", 0)),
        }
        saldo_por_obra.append(item)
        if abs(diff) > 0.05 or (executado_base and not executado_mapa) or (executado_mapa and not executado_base):
            divergencias.append(item)

    divergencias.sort(key=lambda x: abs(x["diferenca_base_mapa"]), reverse=True)
    saldo_por_obra.sort(key=lambda x: x["saldo"], reverse=True)
    total_diff = base["executado"] - mapa["executado_mapa"]
    alertas = []
    if abs(total_diff) > 0.05:
        alertas.append(f"Diferença total entre Base Dash e mapa: {total_diff:,.2f} m")
    if divergencias:
        alertas.append(f"{len(divergencias)} obra(s) com divergência entre executado da Base Dash e mapa.")
    if pds["hoje"] == 0:
        alertas.append("PDS sem registros na data de hoje.")

    return {
        "atualizado_em": now_iso(),
        "totais": {
            "base_planejado": round(base["planejado"], 2),
            "base_executado": round(base["executado"], 2),
            "mapa_executado": round(mapa["executado_mapa"], 2),
            "diferenca_base_mapa": round(total_diff, 2),
            "saldo_total": round(base["planejado"] - base["executado"], 2),
        },
        "pds": pds,
        "alertas": alertas,
        "divergencias": divergencias[:30],
        "saldo_por_obra": saldo_por_obra[:80],
        "arquivos": [
            file_info(BASE_XLSX),
            file_info(MAPA_XLSX),
            file_info(WORKBOOK := DOCS / "pds_word_completo_preenchido(2).xlsx"),
            file_info(PDS_JSON),
        ],
    }


def update_history(payload: dict) -> list:
    hist = []
    if OUT_HIST.exists():
        try:
            hist = json.loads(OUT_HIST.read_text(encoding="utf-8"))
        except Exception:
            hist = []
    entry = {
        "atualizado_em": payload["atualizado_em"],
        "base_executado": payload["totais"]["base_executado"],
        "mapa_executado": payload["totais"]["mapa_executado"],
        "diferenca": payload["totais"]["diferenca_base_mapa"],
        "pds_hoje": payload["pds"]["hoje"],
        "alertas": len(payload["alertas"]),
    }
    hist.append(entry)
    return hist[-200:]


def write_report(payload: dict) -> None:
    totals = payload["totais"]
    lines = [
        "RELATÓRIO DE ATUALIZAÇÃO DO DASHBOARD",
        f"Atualizado em: {payload['atualizado_em']}",
        "",
        f"Planejado Base Dash: {totals['base_planejado']:,.2f} m",
        f"Executado Base Dash: {totals['base_executado']:,.2f} m",
        f"Executado Mapa: {totals['mapa_executado']:,.2f} m",
        f"Diferença Base - Mapa: {totals['diferenca_base_mapa']:,.2f} m",
        f"Saldo total: {totals['saldo_total']:,.2f} m",
        "",
        f"PDS total: {payload['pds']['total']} registros",
        f"PDS hoje: {payload['pds']['hoje']} registros",
        f"Última data PDS: {payload['pds']['ultima_data'] or '-'}",
        "",
        "Alertas:",
    ]
    lines.extend([f"- {a}" for a in payload["alertas"]] or ["- Sem alertas críticos."])
    lines.append("")
    lines.append("Maiores divergências por obra:")
    for item in payload["divergencias"][:10]:
        lines.append(
            f"- {item['obra']}: Base {item['executado_base']:,.2f} m | Mapa {item['executado_mapa']:,.2f} m | Dif. {item['diferenca_base_mapa']:,.2f} m"
        )
    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    payload = build_validation()
    history = update_history(payload)
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_HIST.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
    write_report(payload)

    DOCS.mkdir(exist_ok=True)
    for path in (OUT_JSON, OUT_HIST, OUT_TXT):
        shutil.copy2(path, DOCS / path.name)

    print("Validação automática gerada.")
    print(f"Diferença Base - Mapa: {payload['totais']['diferenca_base_mapa']:,.2f} m")
    print(f"Alertas: {len(payload['alertas'])}")


if __name__ == "__main__":
    main()
