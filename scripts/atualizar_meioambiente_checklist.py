# -*- coding: utf-8 -*-
"""Gera a base do painel de Meio Ambiente a partir do CHECK LIST completo."""

from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "docs" / "CHECK LIST - frente de serviço MAIO.xlsx"
HTML_PATHS = [ROOT / "docs" / "meioambiente.html", ROOT / "meioambiente.html"]


def clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip(" ;")


def key_text(value: str) -> str:
    value = unicodedata.normalize("NFKD", clean(value)).upper()
    return "".join(ch for ch in value if not unicodedata.combining(ch))


def result_from_cell(value) -> tuple[str, str] | None:
    text = clean(value)
    if not text:
        return None
    key = key_text(text)
    if key.startswith("SIM"):
        return "Conforme", text
    if key.startswith("NAO") or key.startswith("NÃO"):
        obs = re.sub(r"^\s*n[ãa]o\s*;?", "", text, flags=re.I).strip(" ;") or "Não"
        return "Não Conforme", obs
    return None


def title_item(header: str) -> str:
    text = clean(header).replace("  ", " ")
    text = text.replace(" - ", " - ")
    return text[:1].upper() + text[1:].lower()


def build_records() -> list[dict]:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    per_leader = defaultdict(lambda: {"Conforme": 0, "Não Conforme": 0})
    per_item = defaultdict(int)

    for row_idx in range(2, ws.max_row + 1):
        dt = ws.cell(row_idx, 1).value
        if not isinstance(dt, datetime):
            continue

        leader = clean(ws.cell(row_idx, 2).value or ws.cell(row_idx, 3).value)
        if not leader:
            leader = "Sem líder"

        month = dt.strftime("%Y-%m-01")

        for col_idx in range(5, ws.max_column + 1):
            parsed = result_from_cell(ws.cell(row_idx, col_idx).value)
            if not parsed:
                continue

            result, obs = parsed
            item = title_item(headers[col_idx - 1])
            per_leader[(month, leader)][result] += 1

            if result == "Não Conforme":
                per_item[(month, item, obs)] += 1

    wb.close()

    records: list[dict] = []
    for (month, leader), counts in sorted(per_leader.items()):
        for result in ("Conforme", "Não Conforme"):
            qty = counts[result]
            if qty:
                records.append(
                    {
                        "data": month,
                        "lider": leader,
                        "item": "Base da planilha",
                        "resultado": result,
                        "observacao": "Base real",
                        "quantidade": qty,
                    }
                )

    for (month, item, obs), qty in sorted(per_item.items()):
        records.append(
            {
                "data": month,
                "lider": "Base Geral",
                "item": item,
                "resultado": "Não Conforme",
                "observacao": obs,
                "quantidade": qty,
            }
        )

    return records


def replace_base(html: str, records: list[dict]) -> str:
    payload = json.dumps(records, ensure_ascii=False, indent=6)
    replacement = "const basePlanilha = " + payload + ";"
    return re.sub(
        r"const basePlanilha = \[[\s\S]*?\];",
        replacement,
        html,
        count=1,
    )


def main() -> None:
    records = build_records()
    months = sorted({record["data"][:7] for record in records})

    for path in HTML_PATHS:
        html = path.read_text(encoding="utf-8-sig")
        updated = replace_base(html, records)
        path.write_text(updated, encoding="utf-8")

    print(f"Meio ambiente atualizado: {len(records)} registros; meses={', '.join(months)}")


if __name__ == "__main__":
    main()
