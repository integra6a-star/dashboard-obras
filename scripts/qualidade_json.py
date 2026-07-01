# -*- coding: utf-8 -*-
"""Gera a base JSON de Qualidade a partir da planilha FOR-Q-05.1."""

from __future__ import annotations

import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
SOURCE_CANDIDATES = [
    ROOT / "FOR-Q-05.1 -  Planilha de indicadores.xlsx",
    ROOT / "docs" / "FOR-Q-05.1 -  Planilha de indicadores.xlsx",
    Path(r"C:\Users\micro\AppData\Local\Temp\FOR-Q-05.1 -  Planilha de indicadores (1).xlsx"),
    Path(r"C:\Users\micro\AppData\Local\Temp\FOR-Q-05.1 -  Planilha de indicadores.xlsx"),
]

MESES = [
    ("janeiro", "Jan"),
    ("fevereiro", "Fev"),
    ("março", "Mar"),
    ("abril", "Abr"),
    ("maio", "Mai"),
    ("junho", "Jun"),
    ("julho", "Jul"),
    ("agosto", "Ago"),
    ("setembro", "Set"),
    ("outubro", "Out"),
    ("novembro", "Nov"),
    ("dezembro", "Dez"),
]


def clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "").replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if match:
        text = match.group(0)
    try:
        return float(text)
    except ValueError:
        return None


def iso_date(value) -> str:
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def meta_value(meta: str, resultado: float | None) -> float | None:
    raw = number(meta)
    if raw is None:
        return None
    if "%" in str(meta):
        return raw / 100
    if resultado is not None and abs(resultado) <= 2 and raw > 2:
        return raw / 100
    return raw


def format_result(value: float | None, target: float | None) -> str:
    if value is None:
        return "-"
    if target is not None and target <= 1.5:
        return f"{value * 100:.1f}%"
    if abs(value) <= 1.5:
        return f"{value * 100:.1f}%"
    return f"{value:.1f}"


def reached(value: float | None, target: float | None) -> bool | None:
    if value is None or target is None:
        return None
    return value >= target


def find_planning(ws) -> dict[str, dict]:
    planning = {}
    for row in range(10, ws.max_row + 1):
        numero = ws.cell(row, 4).value
        indicador = clean(ws.cell(row, 3).value)
        if not indicador:
            continue
        planning[indicador.lower()] = {
            "numero": str(int(numero)) if isinstance(numero, (int, float)) else clean(numero),
            "tema": clean(ws.cell(row, 1).value),
            "objetivo": clean(ws.cell(row, 2).value),
            "calculo": clean(ws.cell(row, 5).value),
            "meta": clean(ws.cell(row, 6).value),
            "frequencia": clean(ws.cell(row, 7).value),
        }
    return planning


def parse_indicadores(wb) -> list[dict]:
    ws = wb["Indicadores"]
    planning = find_planning(wb["Planejamento"]) if "Planejamento" in wb.sheetnames else {}
    indicadores = []
    row = 11
    used_numbers = set()
    while row <= ws.max_row:
        nome = clean(ws.cell(row, 3).value)
        if not nome:
            row += 1
            continue
        resultado_offset = None
        if clean(ws.cell(row + 2, 4).value).lower() == "resultado":
            resultado_offset = 2
        elif clean(ws.cell(row + 1, 4).value).lower() == "resultado":
            resultado_offset = 1
        if resultado_offset is None:
            row += 1
            continue

        plan = planning.get(nome.lower(), {})
        numero = plan.get("numero") or clean(ws.cell(row, 2).value)
        if numero in used_numbers:
            numero = plan.get("numero") or str(len(indicadores) + 1)
        used_numbers.add(numero)

        meta = clean(ws.cell(row, 5).value) or plan.get("meta", "")
        meses = []
        for idx, (mes_key, mes_nome) in enumerate(MESES, start=6):
            numerador = number(ws.cell(row, idx).value)
            denominador = number(ws.cell(row + 1, idx).value) if resultado_offset == 2 else None
            resultado = number(ws.cell(row + resultado_offset, idx).value)
            if numerador is None and denominador is None and resultado is None:
                continue
            if numerador is None and denominador is None and resultado == 0:
                continue
            target = meta_value(meta, resultado)
            meses.append(
                {
                    "mes": mes_key,
                    "mes_nome": mes_nome,
                    "numerador": numerador,
                    "denominador": denominador,
                    "resultado": resultado,
                    "resultado_formatado": format_result(resultado, target),
                    "meta": target,
                    "meta_texto": meta,
                    "atingiu": reached(resultado, target),
                }
            )

        acumulado = {
            "numerador": number(ws.cell(row, 18).value),
            "denominador": number(ws.cell(row + 1, 18).value) if resultado_offset == 2 else None,
            "resultado": number(ws.cell(row + resultado_offset, 18).value),
            "media": number(ws.cell(row + resultado_offset, 19).value),
        }
        target = meta_value(meta, acumulado["resultado"])
        resultado_acumulado = acumulado["resultado"]
        if target is not None and target > 2 and acumulado["media"] is not None:
            resultado_acumulado = acumulado["media"]
        ultimo = next((m for m in reversed(meses) if m["resultado"] is not None), None)
        indicadores.append(
            {
                "numero": numero,
                "nome": nome,
                "tema": plan.get("tema", "QUALIDADE"),
                "objetivo": plan.get("objetivo", ""),
                "calculo": plan.get("calculo", clean(ws.cell(row, 4).value)),
                "meta_texto": meta,
                "meta": target,
                "frequencia": plan.get("frequencia", "Mensal"),
                "medida_1": clean(ws.cell(row, 4).value),
                "medida_2": clean(ws.cell(row + 1, 4).value),
                "meses": meses,
                "acumulado": {
                    **acumulado,
                    "resultado_base": resultado_acumulado,
                    "resultado_formatado": format_result(resultado_acumulado, target),
                    "atingiu": reached(resultado_acumulado, target),
                },
                "ultimo": ultimo,
            }
        )
        row += resultado_offset + 1
    return indicadores


def parse_acoes(wb, indicadores: list[dict]) -> list[dict]:
    if "Plano de ação" not in wb.sheetnames:
        return []
    nomes = {str(item["numero"]): item["nome"] for item in indicadores}
    ws = wb["Plano de ação"]
    acoes = []
    current = ""
    for row in range(10, ws.max_row + 1):
        numero = clean(ws.cell(row, 1).value)
        if numero:
            current = numero
        mes = ws.cell(row, 2).value
        causa = clean(ws.cell(row, 3).value)
        acao = clean(ws.cell(row, 4).value)
        prazo = ws.cell(row, 5).value
        status = clean(ws.cell(row, 6).value) or clean(ws.cell(row, 8).value)
        fechamento = ws.cell(row, 7).value
        if not (causa or acao or status):
            continue
        acoes.append(
            {
                "indicador": current,
                "indicador_nome": nomes.get(current, f"Indicador {current}" if current else "Não informado"),
                "mes": iso_date(mes),
                "causa": causa,
                "acao": acao,
                "prazo": iso_date(prazo),
                "status": status or "Não informado",
                "data_fechamento": iso_date(fechamento),
            }
        )
    return acoes


def build_payload(source: Path) -> dict:
    wb = load_workbook(source, data_only=True)
    indicadores = parse_indicadores(wb)
    acoes = parse_acoes(wb, indicadores)
    total = len(indicadores)
    atingidos = sum(1 for item in indicadores if item["acumulado"].get("atingiu") is True)
    return {
        "metadata": {
            "fonte": source.name,
            "gerado_em": datetime.now().isoformat(timespec="seconds"),
            "estrutura": "FOR-Q-05.1 - Planilha de indicadores",
        },
        "indicadores": indicadores,
        "acoes": acoes,
        "resumo": {
            "total_indicadores": total,
            "indicadores_atingidos": atingidos,
            "indicadores_fora_meta": total - atingidos,
            "acoes_total": len(acoes),
            "acoes_pendentes": sum(1 for a in acoes if a["status"].lower() != "concluído"),
        },
    }


def main() -> None:
    source = next((path for path in SOURCE_CANDIDATES if path.exists()), None)
    if not source:
        raise FileNotFoundError("Planilha de indicadores da Qualidade não encontrada.")
    root_copy = ROOT / "FOR-Q-05.1 -  Planilha de indicadores.xlsx"
    if source.resolve() != root_copy.resolve():
        shutil.copy2(source, root_copy)
    DOCS.mkdir(exist_ok=True)
    docs_copy = DOCS / "FOR-Q-05.1 -  Planilha de indicadores.xlsx"
    if source.resolve() != docs_copy.resolve():
        shutil.copy2(source, docs_copy)
    payload = build_payload(source)
    for target in [ROOT / "qualidade_data.json", DOCS / "qualidade_data.json"]:
        target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        "Qualidade atualizada:",
        f"indicadores={len(payload['indicadores'])}",
        f"acoes={len(payload['acoes'])}",
        f"atingidos={payload['resumo']['indicadores_atingidos']}",
    )


if __name__ == "__main__":
    main()
