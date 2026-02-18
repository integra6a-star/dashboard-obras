import json
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook

ARQ_EXCEL = Path("BASE_DASH_EXTENSAO_POWERBI.xlsx")
SAIDA_JSON = Path("docs/dados.json")

EXPECTED = {
    "Obra": ["obra"],
    "Bloco": ["bloco"],
    "Tipo": ["tipo_extensao", "tipo extensao", "tipo"],
    "Planejado_m": ["extensao_planejada_m", "extensao planejada (m)", "extensao planejada", "planejado_m"],
    "Executado_m": ["extensao_executada_m", "extensao executada (m)", "extensao executada", "executado_m"],
    "PV": ["pv", "pvs", "qtd pv", "quantidade pv"],
    "Profundidade_m": ["profundidade_pv_m", "profundidade pv (m)", "profundidade", "profundidade_m"],
    "Economias_Previstas": ["economias prevista", "economias previstas", "econ prev"],
    "Economias_Recebidas": ["economias recebidas", "economias recebida", "econ receb"],
}

def norm(s):
    return "".join(str(s).strip().lower().split())

def to_float(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except:
        return 0.0

def main():
    if not ARQ_EXCEL.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {ARQ_EXCEL}")

    wb = load_workbook(ARQ_EXCEL, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in ws[1]]
    if not headers or all(h is None for h in headers):
        raise ValueError("Primeira linha da planilha está vazia (sem cabeçalhos).")

    headers_norm = [norm(h) if h is not None else "" for h in headers]

    col_index = {}
    for out_name, candidates in EXPECTED.items():
        found = None
        for cand in candidates:
            cand_n = norm(cand)
            for i, hn in enumerate(headers_norm):
                if hn == cand_n:
                    found = i
                    break
            if found is not None:
                break
        if found is None:
            raise ValueError(f"Não encontrei a coluna para '{out_name}'. Cabeçalhos encontrados: {headers}")
        col_index[out_name] = found

    registros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue

        registros.append({
            "Obra": str(row[col_index["Obra"]] or "").strip(),
            "Bloco": str(row[col_index["Bloco"]] or "").strip(),
            "Tipo": str(row[col_index["Tipo"]] or "").strip(),
            "Planejado_m": to_float(row[col_index["Planejado_m"]]),
            "Executado_m": to_float(row[col_index["Executado_m"]]),
            "PV": to_float(row[col_index["PV"]]),
            "Profundidade_m": to_float(row[col_index["Profundidade_m"]]),
            "Economias_Previstas": to_float(row[col_index["Economias_Previstas"]]),
            "Economias_Recebidas": to_float(row[col_index["Economias_Recebidas"]]),
        })

    payload = {
        "atualizado_em": datetime.now(ZoneInfo("America/Sao_Paulo")).isoformat(),
        "registros": registros
    }

    SAIDA_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(SAIDA_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"OK: docs/dados.json gerado com {len(registros)} linhas.")

if __name__ == "__main__":
    main()
